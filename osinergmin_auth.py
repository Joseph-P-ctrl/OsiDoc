from __future__ import annotations

import argparse
import csv
import logging
import os
import re
import shutil
import sqlite3
import time
import unicodedata
from datetime import datetime, timedelta
from dataclasses import dataclass, replace
from pathlib import Path
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup

try:
    from openpyxl import Workbook  # type: ignore[import-not-found]
except Exception:
    Workbook = None

try:
    from selenium import webdriver
    from selenium.common.exceptions import NoAlertPresentException, TimeoutException, UnexpectedAlertPresentException
    from selenium.webdriver import ActionChains
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.support.ui import WebDriverWait
except Exception:
    webdriver = None
    TimeoutException = Exception
    UnexpectedAlertPresentException = Exception
    NoAlertPresentException = Exception
    ActionChains = None
    By = None
    Keys = None
    EC = None
    WebDriverWait = None


@dataclass
class AuthConfig:
    login_url: str = "https://osivirtual.osinergmin.gob.pe/autenticacion/acceso-sistema"
    base_url: str = "https://osivirtual.osinergmin.gob.pe"
    username_id: str = "documentoIdentidad"
    password_id: str = "contrasena"
    login_form_selector: str = "form"
    login_submit_selector: str = "button[type='submit']"
    post_login_selector: str = ""
    login_error_selector: str = ""
    success_url_keyword: str = "autenticacion/acceso-sistema"
    selenium_headless: bool = True
    timeout: int = 30
    captcha_timeout: int = 180
    open_sne_after_login: bool = True
    require_sne_click_navigation: bool = True
    sne_menu_selector: str = (
        "//div[@matlistitemtitle and contains(@class,'text-menu-parent') "
        "and contains(normalize-space(),'Casilla') and contains(normalize-space(),'SNE')]"
    )
    sne_target_url: str = "https://notificaciones.osinergmin.gob.pe/sne-web/pages/notificacion/inicio"
    sne_expected_text: str = "Sistema de Notificaciones Electrónicas|Bandeja de Entrada"
    fecha_notificacion_inicio: str = ""
    fecha_notificacion_fin: str = ""
    sne_fecha_inicio_id: str = "fechaNotificacionInicio"
    sne_fecha_fin_id: str = "fechaNotificacionFin"
    sne_leido_value: str = ""
    sne_buscar_button_id: str = "buscar-boton"
    sne_export_excel_selector: str = (
        "//div[contains(@class,'ui-pg-div') and .//span[contains(@class,'ui-icon-arrowthickstop-1-s')] "
        "and contains(normalize-space(.), 'Exportar a Excel')]"
    )
    download_dir: str = "downloads"
    export_wait_seconds: int = 40
    target_notifications: tuple[str, ...] = ()
    incremental_only: bool = True
    skip_existing_notifications: bool = True
    user_agent: str = (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )


def _parse_ddmmyyyy(raw: str) -> datetime | None:
    if not raw:
        return None
    try:
        return datetime.strptime(raw.strip(), "%d/%m/%Y")
    except ValueError:
        return None


def _parse_iso_date(raw: str) -> datetime | None:
    if not raw:
        return None
    try:
        return datetime.strptime(raw.strip(), "%Y-%m-%d")
    except ValueError:
        return None


def _resolve_processing_date(cfg: AuthConfig) -> str:
    """Fecha del lote procesado en formato YYYY-MM-DD (usa fecha_fin del filtro)."""
    by_filter = _parse_ddmmyyyy(cfg.fecha_notificacion_fin)
    if by_filter is not None:
        return by_filter.strftime("%Y-%m-%d")
    return datetime.now().strftime("%Y-%m-%d")


def _env_bool(name: str, default: bool) -> bool:
    raw = os.getenv(name)
    if raw is None:
        return default
    return raw.strip().lower() in {"1", "true", "yes", "si", "on"}


def _parse_target_notifications(raw: str) -> tuple[str, ...]:
    """Parsea lista CSV/espaciada de Nro. Notificacion y devuelve valores unicos."""
    if not raw:
        return ()

    parts = [p.strip() for p in re.split(r"[\s,;]+", raw) if p.strip()]
    out: list[str] = []
    seen: set[str] = set()
    for part in parts:
        val = re.sub(r"[^0-9\-]", "", part).strip("-")
        if not re.fullmatch(r"\d{8,}-\d+", val):
            continue
        if val in seen:
            continue
        seen.add(val)
        out.append(val)
    return tuple(out)


def _load_dotenv(dotenv_path: str = ".env") -> None:
    """Carga variables de entorno desde un archivo .env sin dependencias externas."""
    path = Path(dotenv_path)
    if not path.exists():
        return

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        raw_line = raw_line.lstrip("\ufeff")
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue

        key, value = line.split("=", 1)
        key = key.strip().lstrip("\ufeff")
        value = value.strip().strip('"').strip("'")
        if not key:
            continue

        # Las variables OSI_* del proyecto deben sobrescribir valores heredados del sistema.
        if key.startswith("OSI_"):
            os.environ[key] = value
        elif os.getenv(key) is None:
            os.environ[key] = value


def _extract_hidden_inputs(form) -> dict[str, str]:
    payload: dict[str, str] = {}
    if not form:
        return payload

    for hidden in form.select("input[type='hidden']"):
        name = hidden.get("name")
        value = hidden.get("value", "")
        if name:
            payload[name] = value

    return payload


def _is_authenticated_by_url(url: str, success_url_keyword: str) -> bool:
    # Si seguimos en acceso-sistema, asumimos que no se autenticó.
    return success_url_keyword.lower() not in url.lower()


def _resolve_input_name_by_id(form, input_id: str, fallback_name: str) -> str:
    if not form:
        return fallback_name

    input_el = form.select_one(f"#{input_id}")
    if not input_el:
        return fallback_name

    return input_el.get("name") or fallback_name


def _login_with_requests(session: requests.Session, cfg: AuthConfig, username: str, password: str) -> bool:
    try:
        login_page = session.get(cfg.login_url, timeout=cfg.timeout)
        login_page.raise_for_status()
    except requests.RequestException as exc:
        raise RuntimeError(f"Error de conexion al abrir login: {exc}") from exc

    soup = BeautifulSoup(login_page.text, "html.parser")
    form = soup.select_one(cfg.login_form_selector)
    if not form:
        return False

    payload = _extract_hidden_inputs(form)
    username_key = _resolve_input_name_by_id(form, cfg.username_id, cfg.username_id)
    password_key = _resolve_input_name_by_id(form, cfg.password_id, cfg.password_id)
    payload[username_key] = username
    payload[password_key] = password

    post_url = cfg.login_url
    if form.get("action"):
        post_url = urljoin(cfg.base_url, form.get("action"))

    try:
        response = session.post(post_url, data=payload, timeout=cfg.timeout, allow_redirects=True)
        response.raise_for_status()
    except requests.RequestException as exc:
        raise RuntimeError(f"Error al enviar credenciales: {exc}") from exc

    final_url = response.url or cfg.login_url
    if _is_authenticated_by_url(final_url, cfg.success_url_keyword):
        return True

    # Verificacion adicional navegando nuevamente a login.
    check = session.get(cfg.login_url, timeout=cfg.timeout, allow_redirects=True)
    check.raise_for_status()
    return _is_authenticated_by_url(check.url, cfg.success_url_keyword)


def _new_driver(cfg: AuthConfig):
    if webdriver is None:
        raise RuntimeError("Selenium no esta disponible en el entorno.")

    download_path = Path(cfg.download_dir).expanduser().resolve()
    download_path.mkdir(parents=True, exist_ok=True)

    options = webdriver.ChromeOptions()
    if cfg.selenium_headless:
        options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_experimental_option(
        "prefs",
        {
            "download.default_directory": str(download_path),
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True,
            # Evita el popup "Descargar varios archivos" (Permitir/Bloquear).
            "profile.default_content_setting_values.automatic_downloads": 1,
        },
    )
    driver = webdriver.Chrome(options=options)
    try:
        driver.execute_cdp_cmd(
            "Browser.setDownloadBehavior",
            {
                "behavior": "allow",
                "downloadPath": str(download_path),
                "eventsEnabled": False,
            },
        )
    except Exception:
        # No bloquea el flujo si esta capacidad no esta disponible.
        pass
    return driver


def _find_first(driver, selectors: list[tuple[str, str]], timeout: int):
    for by, value in selectors:
        try:
            return WebDriverWait(driver, timeout).until(EC.presence_of_element_located((by, value)))
        except Exception:
            continue
    return None


def _find_ingresar_button(driver, cfg: AuthConfig, timeout: int):
    # Prioriza boton/input cuyo texto o value sea exactamente 'Ingresar'.
    return _find_first(
        driver,
        [
            (By.XPATH, "//button[contains(translate(normalize-space(.), 'INGRESAR', 'ingresar'), 'ingresar') ]"),
            (By.XPATH, "//button[contains(@class,'bg-primary') and contains(translate(normalize-space(.), 'INGRESAR', 'ingresar'), 'ingresar') ]"),
            (By.CSS_SELECTOR, "input[type='submit'][value='Ingresar']"),
            (By.XPATH, "//button[normalize-space()='Ingresar']"),
            (By.XPATH, "//input[@type='submit' and @value='Ingresar']"),
            (By.CSS_SELECTOR, cfg.login_submit_selector),
            (By.CSS_SELECTOR, "button[type='submit']"),
            (By.CSS_SELECTOR, "button"),
            (By.CSS_SELECTOR, "input[type='submit']"),
        ],
        timeout,
    )


def _click_ingresar_fallback(driver, password_id: str) -> bool:
    """Fallback: clic por texto 'Ingresar' via JS o envio con Enter."""
    script = """
const pwdId = arguments[0];
const nodes = Array.from(document.querySelectorAll('button, input[type="submit"]'));
const btn = nodes.find((n) => {
    const raw = (n.innerText || n.textContent || n.value || '').trim().toLowerCase();
    return raw.includes('ingresar');
});

if (btn) {
    btn.click();
    return true;
}

const pwd = document.getElementById(pwdId);
if (pwd) {
    pwd.focus();
    const ev = new KeyboardEvent('keydown', { key: 'Enter', code: 'Enter', keyCode: 13, which: 13, bubbles: true });
    pwd.dispatchEvent(ev);
    return true;
}

return false;
"""
    try:
        return bool(driver.execute_script(script, password_id))
    except Exception:
        return False


def _click_ingresar_button(driver, submit) -> bool:
    """Hace clic explícito en Ingresar con múltiples estrategias."""
    if submit is None:
        return False

    # Lleva el botón al viewport antes del clic.
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", submit)
    except Exception:
        pass

    # Estrategia 1: click directo Selenium.
    try:
        submit.click()
        return True
    except Exception:
        pass

    # Estrategia 2: ActionChains (algunos overlays lo requieren).
    try:
        if ActionChains is not None:
            ActionChains(driver).move_to_element(submit).click().perform()
            return True
    except Exception:
        pass

    # Estrategia 3: disparo explícito de eventos de mouse + click JS.
    try:
        driver.execute_script(
            """
const el = arguments[0];
['mouseenter','mouseover','mousemove','mousedown','mouseup','click'].forEach((type) => {
  const evt = new MouseEvent(type, { bubbles: true, cancelable: true, view: window });
  el.dispatchEvent(evt);
});
""",
            submit,
        )
        return True
    except Exception:
        return False


def _click_ingresar_with_retries(driver, cfg: AuthConfig, retries: int = 6, delay_seconds: float = 1.0) -> bool:
    """Intenta hacer clic en Ingresar varias veces por cambios dinámicos del DOM."""
    for intento in range(1, retries + 1):
        submit = _find_ingresar_button(driver, cfg, 2)
        if submit is not None:
            logging.info("Intento %s: clic en boton Ingresar por selector.", intento)
            if _click_ingresar_button(driver, submit):
                return True

        logging.info("Intento %s: fallback JS/Enter para enviar login.", intento)
        if _click_ingresar_fallback(driver, cfg.password_id):
            return True

        time.sleep(delay_seconds)

    return False


def _normalize_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value or "")
    return "".join(ch for ch in normalized if not unicodedata.combining(ch)).lower()


def _wait_for_sne_home(driver, cfg: AuthConfig) -> bool:
    """Valida que estamos en la pagina SNE esperada por URL y texto visible."""
    expected_tokens = [_normalize_text(part.strip()) for part in cfg.sne_expected_text.split("|") if part.strip()]
    expected_url = _normalize_text(cfg.sne_target_url)

    def _condition(drv) -> bool:
        current_url = _normalize_text(drv.current_url or "")

        # Se exige estar en el dominio/ruta objetivo del SNE.
        if expected_url and expected_url not in current_url:
            return False

        try:
            page_text = drv.execute_script(
                "return (document.body && document.body.innerText) ? document.body.innerText : '';"
            )
        except Exception:
            page_text = drv.page_source or ""

        page = _normalize_text(page_text)

        # Se valida el contenido principal que aparece en la vista esperada.
        if expected_tokens and not all(token in page for token in expected_tokens):
            return False

        return True

    try:
        WebDriverWait(driver, cfg.timeout).until(_condition)
        return True
    except TimeoutException:
        return False


def _set_input_value(driver, element, value: str) -> None:
    try:
        element.clear()
    except Exception:
        pass

    try:
        element.send_keys(Keys.CONTROL, "a")
        element.send_keys(Keys.DELETE)
    except Exception:
        pass

    try:
        element.send_keys(value)
    except Exception:
        pass

    # Dispara eventos necesarios para frameworks como jQuery datepicker.
    try:
        driver.execute_script(
            """
const el = arguments[0];
const val = arguments[1];
if (el.value !== val) { el.value = val; }
['input','change','blur'].forEach((t) => {
    el.dispatchEvent(new Event(t, { bubbles: true }));
});
""",
            element,
            value,
        )
    except Exception:
        pass


def _accept_browser_alert_if_present(driver, expected_text: str = "") -> str:
    """Acepta alertas del navegador para evitar que rompan el flujo Selenium."""
    try:
        alert = driver.switch_to.alert
        text = (alert.text or "").strip()
        alert.accept()

        if expected_text and expected_text.lower() in text.lower():
            logging.info("Alerta esperada detectada y aceptada: %s", text)
        elif text:
            logging.warning("Alerta detectada y aceptada: %s", text)

        return text
    except NoAlertPresentException:
        return ""
    except Exception:
        return ""


def _snapshot_downloads(download_dir: Path) -> dict[str, float]:
    snapshot: dict[str, float] = {}
    if not download_dir.exists():
        return snapshot

    for file_path in download_dir.iterdir():
        if not file_path.is_file():
            continue
        try:
            snapshot[file_path.name] = file_path.stat().st_mtime
        except OSError:
            continue
    return snapshot


def _wait_for_new_download(download_dir: Path, before: dict[str, float], timeout_seconds: int) -> Path | None:
    end_time = time.time() + timeout_seconds
    temp_suffixes = (".crdownload", ".tmp", ".part")

    while time.time() < end_time:
        if not download_dir.exists():
            time.sleep(0.25)
            continue

        candidates: list[Path] = []
        for file_path in download_dir.iterdir():
            if not file_path.is_file():
                continue
            if file_path.name.lower().endswith(temp_suffixes):
                continue
            candidates.append(file_path)

        candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        for file_path in candidates:
            try:
                current_mtime = file_path.stat().st_mtime
            except OSError:
                continue

            previous_mtime = before.get(file_path.name)
            if previous_mtime is None or current_mtime > previous_mtime:
                return file_path

        time.sleep(0.5)

    return None


def _get_sne_grid_snapshot(driver) -> tuple[str, tuple[str, ...]]:
    """Captura un estado simple del grid para detectar si la busqueda realmente refresco."""
    try:
        snapshot = driver.execute_script(
            """
const pagerCandidates = Array.from(document.querySelectorAll("td[id$='_right'], .ui-paging-info"));
const pagerText = pagerCandidates.map((el) => (el.textContent || '').trim()).find(Boolean) || '';
const isVisible = (el) => {
    if (!el) return false;
    const style = window.getComputedStyle(el);
    return style.display !== 'none' && style.visibility !== 'hidden' && el.offsetParent !== null;
};
const rows = Array.from(document.querySelectorAll('table.ui-jqgrid-btable tr.jqgrow'))
        .filter((row) => isVisible(row))
    .map((row) => (row.textContent || '').replace(/\\s+/g, ' ').trim())
    .filter(Boolean)
  .filter(Boolean)
  .slice(0, 5);
return { pagerText, rows };
"""
        )
        pager_text = (snapshot or {}).get("pagerText", "")
        rows = tuple((snapshot or {}).get("rows", []))
        return pager_text, rows
    except Exception:
        return "", tuple()


def _get_sne_grid_runtime_state(driver) -> tuple[bool, str, int]:
        """Devuelve estado de carga y un resumen del grid para confirmar refresco."""
        try:
                state = driver.execute_script(
                        """
const isVisible = (el) => {
    if (!el) return false;
    const style = window.getComputedStyle(el);
    return style.display !== 'none' && style.visibility !== 'hidden' && Number(style.opacity || '1') > 0;
};

const loadingEls = Array.from(document.querySelectorAll("div[id^='load_'], div[id^='lui_'], .loading"));
const loadingVisible = loadingEls.some((el) => isVisible(el));
const pagerText = (Array.from(document.querySelectorAll("td[id$='_right'], .ui-paging-info"))
    .map((el) => (el.textContent || '').trim())
    .find(Boolean)) || '';
const rowCount = Array.from(document.querySelectorAll('table.ui-jqgrid-btable tr.jqgrow'))
        .filter((row) => isVisible(row))
        .length;
return { loadingVisible, pagerText, rowCount };
"""
                )
                return bool((state or {}).get("loadingVisible")), str((state or {}).get("pagerText", "")), int((state or {}).get("rowCount", 0))
        except Exception:
                return False, "", 0


def _click_search_button_and_wait(driver, buscar, cfg: AuthConfig) -> bool:
    """Hace clic real en Buscar y espera a que el listado del SNE se refresque."""
    before_snapshot = _get_sne_grid_snapshot(driver)

    def _js_click_buscar(fallback):
        try:
            return bool(driver.execute_script(
                """
const fallback = arguments[0];
const target = document.getElementById('buscar-boton')
    || Array.from(document.querySelectorAll("input[type='button'], button"))
        .find((node) => ((node.value || node.textContent || '').trim() === 'Buscar'))
    || fallback;

if (!target) return false;

try { target.scrollIntoView({ block: 'center' }); } catch (e) {}
try { target.focus(); } catch (e) {}

['pointerdown','mousedown','pointerup','mouseup','click'].forEach((type) => {
    target.dispatchEvent(new MouseEvent(type, { bubbles: true, cancelable: true, view: window }));
});

if (window.jQuery) {
    try { window.jQuery(target).trigger('click'); } catch (e) {}
}
try {
    if (typeof target.onclick === 'function') { target.onclick(); }
} catch (e) {}

try { target.click(); } catch (e) {}
return true;
""",
                fallback,
            ))
        except Exception:
            return False

    for _ in range(2):
        clicked = _js_click_buscar(buscar) or _click_ingresar_button(driver, buscar)
        if not clicked:
            continue

        loading_seen = False
        end_time = time.time() + cfg.timeout
        while time.time() < end_time:
            loading_visible, pager_text, row_count = _get_sne_grid_runtime_state(driver)
            if loading_visible:
                loading_seen = True
            if loading_seen and not loading_visible:
                return True

            after_snapshot = _get_sne_grid_snapshot(driver)
            if after_snapshot != before_snapshot and any(after_snapshot):
                return True

            # Si el grid ya tiene estado legible, damos por aceptada la recarga.
            if pager_text and row_count >= 0:
                if loading_seen:
                    return True

            time.sleep(0.25)

    return False


def _click_export_excel_button(driver, export_btn) -> bool:
    """Hace clic sobre el div exacto de Exportar a Excel con una ruta JS mas robusta."""
    if export_btn is None:
        return False

    if _click_ingresar_button(driver, export_btn):
        return True

    try:
        clicked = driver.execute_script(
            """
const fallback = arguments[0];
const nodes = Array.from(document.querySelectorAll('div.ui-pg-div'));
const target = nodes.find((node) => {
    const text = (node.textContent || '').replace(/\u00a0/g, ' ').replace(/\\s+/g, ' ').trim();
    return text.includes('Exportar a Excel') && node.querySelector('span.ui-icon.ui-icon-arrowthickstop-1-s');
}) || fallback;

if (!target) return false;

try { target.scrollIntoView({ block: 'center' }); } catch (e) {}
try { target.focus(); } catch (e) {}

['pointerdown','mousedown','pointerup','mouseup','click'].forEach((type) => {
    target.dispatchEvent(new MouseEvent(type, { bubbles: true, cancelable: true, view: window }));
});

try { target.click(); } catch (e) {}
return true;
""",
            export_btn,
        )
        return bool(clicked)
    except Exception:
        return False


def _collect_filtered_grid_rows(driver, cfg: AuthConfig) -> tuple[list[str], list[list[str]]]:
    """Recorre el jqGrid paginado y devuelve todas las filas del resultado filtrado."""
    headers: list[str] = []
    all_rows: list[list[str]] = []
    seen_pages: set[str] = set()

    def _read_page() -> tuple[str, list[str], list[list[str]]]:
        try:
            snapshot = driver.execute_script(
                """
const clean = (v) => (v || '').replace(/\u00a0/g, ' ').replace(/\\s+/g, ' ').trim();
const isVisible = (el) => {
    if (!el) return false;
    const style = window.getComputedStyle(el);
    return style.display !== 'none' && style.visibility !== 'hidden' && el.offsetParent !== null;
};
const pagerText = (Array.from(document.querySelectorAll("td[id$='_right'], .ui-paging-info"))
    .map((el) => clean(el.textContent))
    .find(Boolean)) || '';

const headerRow = document.querySelector('table.ui-jqgrid-htable tr.ui-jqgrid-labels');
const headers = headerRow
    ? Array.from(headerRow.querySelectorAll('th')).map((th) => clean(th.textContent))
    : [];

const rows = Array.from(document.querySelectorAll('table.ui-jqgrid-btable tr.jqgrow'))
    .filter((tr) => isVisible(tr))
    .map((tr) => Array.from(tr.querySelectorAll('td')).map((td) => clean(td.textContent)))
    .filter((cells) => cells.some((v) => v));

return { pagerText, headers, rows };
"""
            )
            return str((snapshot or {}).get("pagerText", "")), list((snapshot or {}).get("headers", [])), list((snapshot or {}).get("rows", []))
        except Exception:
            return "", [], []

    def _click_next_page() -> bool:
        try:
            result = driver.execute_script(
                """
const buttons = Array.from(document.querySelectorAll('td.ui-pg-button'));
const nextBtn = buttons.find((td) => /_next$/i.test(td.id || '') || td.querySelector('.ui-icon-seek-next'));
if (!nextBtn) return false;
if (nextBtn.classList.contains('ui-state-disabled') || (nextBtn.getAttribute('aria-disabled') || '') === 'true') {
    return false;
}
try { nextBtn.scrollIntoView({ block: 'center' }); } catch (e) {}
['pointerdown','mousedown','pointerup','mouseup','click'].forEach((type) => {
    nextBtn.dispatchEvent(new MouseEvent(type, { bubbles: true, cancelable: true, view: window }));
});
try { nextBtn.click(); } catch (e) {}
return true;
"""
            )
            return bool(result)
        except Exception:
            return False

    for _ in range(200):
        pager_text, page_headers, page_rows = _read_page()
        page_key = pager_text or f"rows:{len(page_rows)}:{len(all_rows)}"
        if page_key in seen_pages:
            break
        seen_pages.add(page_key)

        if page_headers and not headers:
            headers = page_headers
        if page_rows:
            all_rows.extend(page_rows)

        if not _click_next_page():
            break

        end_time = time.time() + cfg.timeout
        while time.time() < end_time:
            new_pager_text, _, _ = _read_page()
            if new_pager_text != pager_text:
                break
            time.sleep(0.2)

    if not headers and all_rows:
        headers = [f"Columna {i + 1}" for i in range(len(all_rows[0]))]

    # Elimina columnas completamente vacias (ej. checkbox/acciones) para un Excel limpio.
    if headers and all_rows:
        keep_idx: list[int] = []
        col_count = min(len(headers), min(len(r) for r in all_rows))
        for i in range(col_count):
            header = (headers[i] or "").strip()
            has_value = any((r[i] or "").strip() for r in all_rows)
            if header or has_value:
                keep_idx.append(i)

        if keep_idx:
            headers = [headers[i] for i in keep_idx]
            all_rows = [[row[i] for i in keep_idx] for row in all_rows]

    return headers, all_rows


def _get_visible_lupa_count(driver) -> int:
    """Cuenta lupitas visibles de Lectura de Notificacion en el grid actual."""
    try:
        return int(
            driver.execute_script(
                """
const isVisible = (el) => {
    if (!el) return false;
    const style = window.getComputedStyle(el);
    return style.display !== 'none' && style.visibility !== 'hidden' && el.offsetParent !== null;
};
const rows = Array.from(document.querySelectorAll('table.ui-jqgrid-btable tr.jqgrow')).filter((tr) => isVisible(tr));
let count = 0;
for (const row of rows) {
    const icon = row.querySelector("img[title*='Lectura'], img[src*='icon_search.png']");
    if (icon && isVisible(icon)) count += 1;
}
return count;
"""
            )
        )
    except Exception:
        return 0


def _click_lupa_by_index(driver, index: int) -> bool:
    """Hace clic en la lupita de una fila visible por indice (0-based)."""
    try:
        return bool(
            driver.execute_script(
                """
const idx = arguments[0];
const isVisible = (el) => {
    if (!el) return false;
    const style = window.getComputedStyle(el);
    return style.display !== 'none' && style.visibility !== 'hidden' && el.offsetParent !== null;
};
const rows = Array.from(document.querySelectorAll('table.ui-jqgrid-btable tr.jqgrow')).filter((tr) => isVisible(tr));
if (idx < 0 || idx >= rows.length) return false;
const icon = rows[idx].querySelector("img[title*='Lectura'], img[src*='icon_search.png']");
if (!icon || !isVisible(icon)) return false;
try { icon.scrollIntoView({ block: 'center' }); } catch (e) {}
['pointerdown','mousedown','pointerup','mouseup','click'].forEach((t) => {
    icon.dispatchEvent(new MouseEvent(t, { bubbles: true, cancelable: true, view: window }));
});
try { icon.click(); } catch (e) {}
return true;
""",
                index,
            )
        )
    except Exception:
        return False


def _click_lupa_by_notification_number(driver, notification_number: str) -> bool:
    """Hace clic en la lupita buscando la fila por numero de notificacion."""
    if not notification_number:
        return False

    try:
        return bool(
            driver.execute_script(
                """
const wanted = arguments[0];
const isVisible = (el) => {
    if (!el) return false;
    const style = window.getComputedStyle(el);
    return style.display !== 'none' && style.visibility !== 'hidden' && el.offsetParent !== null;
};
const rows = Array.from(document.querySelectorAll('table.ui-jqgrid-btable tr.jqgrow')).filter((tr) => isVisible(tr));

for (const row of rows) {
    const cells = Array.from(row.querySelectorAll('td')).map((td) => (td.textContent || '').trim());
    const rowText = (row.textContent || '').replace(/\\s+/g, ' ').trim();
    const hasNotif = cells.some((v) => v === wanted) || rowText.includes(wanted);
    if (!hasNotif) continue;

    const icon = row.querySelector("img[title*='Lectura'], img[src*='icon_search.png']");
    if (!icon || !isVisible(icon)) return false;
    try { icon.scrollIntoView({ block: 'center' }); } catch (e) {}
    ['pointerdown','mousedown','pointerup','mouseup','click'].forEach((t) => {
        icon.dispatchEvent(new MouseEvent(t, { bubbles: true, cancelable: true, view: window }));
    });
    try { icon.click(); } catch (e) {}
    return true;
}

return false;
""",
                notification_number,
            )
        )
    except Exception:
        return False


def _go_to_next_grid_page(driver, cfg: AuthConfig) -> bool:
    """Avanza a la siguiente pagina del jqGrid si el pager Next esta habilitado."""
    try:
        _, pager_before, _, _ = _get_visible_grid_status(driver)
        clicked = bool(
            driver.execute_script(
                """
const isVisible = (el) => {
    if (!el) return false;
    const style = window.getComputedStyle(el);
    return style.display !== 'none' && style.visibility !== 'hidden' && el.offsetParent !== null;
};

const nextCandidates = Array.from(document.querySelectorAll(
    "span.ui-icon.ui-icon-seek-next, .ui-pg-button .ui-icon-seek-next, .ui-icon-seek-next, td[id^='next_']"
));

for (const node of nextCandidates) {
    const icon = node.matches("span.ui-icon.ui-icon-seek-next") ? node : node.querySelector?.("span.ui-icon.ui-icon-seek-next");
    const target = icon || (node.matches("td") ? node : (node.closest("td") || node));
    const clickable = target.matches("td") ? target : (target.closest("td") || target);
    if (!clickable || !isVisible(clickable)) continue;
    const cls = (target.className || '').toLowerCase();
    const tdCls = (clickable.className || '').toLowerCase();
    if (cls.includes('disabled') || cls.includes('ui-state-disabled') || tdCls.includes('disabled') || tdCls.includes('ui-state-disabled')) continue;
    try { clickable.scrollIntoView({ block: 'center' }); } catch (e) {}
    try { target.click(); return true; } catch (e) {}
    try { clickable.click(); return true; } catch (e) {}
    try {
        ['pointerdown','mousedown','pointerup','mouseup','click'].forEach((t) => {
            target.dispatchEvent(new MouseEvent(t, { bubbles: true, cancelable: true, view: window }));
        });
        return true;
    } catch (e) {}
}

return false;
"""
            )
        )
        if not clicked:
            return False

        # Espera a que el grid refleje cambio de pagina.
        end_time = time.time() + max(2.0, cfg.timeout / 2)
        while time.time() < end_time:
            _, pager_now, _, _ = _get_visible_grid_status(driver)
            if (pager_now or "") != (pager_before or ""):
                time.sleep(0.2)
                return True
            time.sleep(0.15)
        return False
    except Exception:
        return False


def _go_to_first_grid_page(driver, cfg: AuthConfig) -> bool:
    """Regresa a la primera pagina del jqGrid si el pager First esta habilitado."""
    try:
        moved = bool(
            driver.execute_script(
                """
const isVisible = (el) => {
    if (!el) return false;
    const style = window.getComputedStyle(el);
    return style.display !== 'none' && style.visibility !== 'hidden' && el.offsetParent !== null;
};

const firstCandidates = Array.from(document.querySelectorAll(
    "td[id^='first_'], .ui-pg-button .ui-icon-seek-first, .ui-icon-seek-first"
));

for (const node of firstCandidates) {
    const target = node.matches("td") ? node : (node.closest("td") || node);
    if (!target || !isVisible(target)) continue;
    const cls = (target.className || '').toLowerCase();
    if (cls.includes('disabled') || cls.includes('ui-state-disabled')) {
        return false;
    }
    try { target.scrollIntoView({ block: 'center' }); } catch (e) {}
    try { target.click(); return true; } catch (e) {}
    try {
        ['pointerdown','mousedown','pointerup','mouseup','click'].forEach((t) => {
            target.dispatchEvent(new MouseEvent(t, { bubbles: true, cancelable: true, view: window }));
        });
        return true;
    } catch (e) {}
}

return false;
"""
            )
        )

        if moved:
            end_time = time.time() + max(2.0, cfg.timeout / 2)
            while time.time() < end_time:
                _get_visible_grid_status(driver)
                time.sleep(0.2)
                return True
        return True
    except Exception:
        return False


def _click_lupa_by_notification_number_any_page(driver, cfg: AuthConfig, notification_number: str) -> bool:
    """Busca la notificacion en la pagina actual y siguientes paginas del grid."""
    if not notification_number:
        return False

    # Siempre vuelve al inicio para no perder notificaciones que quedaron en paginas anteriores.
    _go_to_first_grid_page(driver, cfg)

    # Intenta primero en la primera pagina.
    if _click_lupa_by_notification_number(driver, notification_number):
        return True

    # Si no esta visible, avanza por paginas y vuelve a intentar.
    max_page_hops = 120
    for _ in range(max_page_hops):
        if not _go_to_next_grid_page(driver, cfg):
            break
        if _click_lupa_by_notification_number(driver, notification_number):
            return True

    return False


def _click_documentos_notificados(driver, cfg: AuthConfig) -> bool:
    """Abre la seccion Documentos notificados dentro del detalle de notificacion."""
    link = _find_first(
        driver,
        [
            (By.ID, "verDocumentosNotificacion-link"),
            (By.XPATH, "//a[@id='verDocumentosNotificacion-link' and contains(normalize-space(.), 'Documentos notificados') ]"),
            (By.XPATH, "//a[contains(normalize-space(.), 'Documentos notificados')]"),
        ],
        cfg.timeout,
    )
    if link is None:
        return False

    try:
        if _click_ingresar_button(driver, link):
            return True
    except Exception:
        pass

    try:
        driver.execute_script("arguments[0].click();", link)
        return True
    except Exception:
        return False


def _download_visible_document_links(driver, cfg: AuthConfig, download_dir: Path) -> int:
    """Descarga todos los links visibles 'Descargar archivo' del detalle abierto."""
    downloaded = 0

    def _count_links() -> int:
        try:
            return int(
                driver.execute_script(
                    """
const isVisible = (el) => {
    if (!el) return false;
    const style = window.getComputedStyle(el);
    return style.display !== 'none' && style.visibility !== 'hidden' && el.offsetParent !== null;
};
const links = Array.from(document.querySelectorAll("a[title='Descargar archivo'], a[href*='descargarArchivoNotificacion']"))
    .filter((a) => isVisible(a));
return links.length;
"""
                )
            )
        except Exception:
            return 0

    link_count = _count_links()
    for idx in range(link_count):
        before = _snapshot_downloads(download_dir)
        try:
            clicked = bool(
                driver.execute_script(
                    """
const index = arguments[0];
const isVisible = (el) => {
    if (!el) return false;
    const style = window.getComputedStyle(el);
    return style.display !== 'none' && style.visibility !== 'hidden' && el.offsetParent !== null;
};
const links = Array.from(document.querySelectorAll("a[title='Descargar archivo'], a[href*='descargarArchivoNotificacion']"))
    .filter((a) => isVisible(a));
if (index < 0 || index >= links.length) return false;
const link = links[index];
try { link.scrollIntoView({ block: 'center' }); } catch (e) {}
['pointerdown','mousedown','pointerup','mouseup','click'].forEach((t) => {
    link.dispatchEvent(new MouseEvent(t, { bubbles: true, cancelable: true, view: window }));
});
try { link.click(); } catch (e) {}
return true;
""",
                    idx,
                )
            )
        except Exception:
            clicked = False

        if not clicked:
            continue

        _accept_browser_alert_if_present(driver)
        file_path = _wait_for_new_download(download_dir, before, max(10, cfg.export_wait_seconds))
        if file_path is not None:
            downloaded += 1

    return downloaded


def _get_notification_number_by_index(driver, index: int) -> str:
    """Obtiene el numero de notificacion de una fila visible por indice."""
    try:
        raw_value = str(
            driver.execute_script(
                """
const idx = arguments[0];
const isVisible = (el) => {
    if (!el) return false;
    const style = window.getComputedStyle(el);
    return style.display !== 'none' && style.visibility !== 'hidden' && el.offsetParent !== null;
};
const rows = Array.from(document.querySelectorAll('table.ui-jqgrid-btable tr.jqgrow')).filter((tr) => isVisible(tr));
if (idx < 0 || idx >= rows.length) return '';
const cells = Array.from(rows[idx].querySelectorAll('td')).map((td) => (td.textContent || '').trim());
for (const cell of cells) {
    if (/\\d{8,}-\\d+/.test(cell)) return cell;
}
const rowText = (rows[idx].textContent || '').replace(/\\s+/g, ' ').trim();
const match = rowText.match(/\\d{8,}-\\d+/);
return match ? match[0] : '';
""",
                index,
            )
        )
    except Exception:
        raw_value = ""

    clean = re.sub(r"[^0-9\-]", "", raw_value or "").strip("-")
    return clean


def _move_download_to_notification_folder(
    file_path: Path,
    base_download_dir: Path,
    notification_number: str,
    processing_date: str,
) -> Path:
    """Mueve un archivo descargado a /YYYY-MM-DD/numero_suministro/."""
    folder_name = notification_number or "sin-numero-suministro"
    date_folder = processing_date
    target_dir = base_download_dir / date_folder / folder_name
    target_dir.mkdir(parents=True, exist_ok=True)

    target_path = target_dir / file_path.name
    if target_path.exists():
        try:
            target_path.unlink()
        except Exception:
            pass

    try:
        file_path.replace(target_path)
    except Exception:
        return file_path
    return target_path


def _get_notifications_with_downloads(base_download_dir: Path, processing_date: str) -> set[str]:
    """Obtiene Nro. Notificacion que ya tienen archivos descargados en CUALQUIER carpeta de fecha.
    Revisa todas las subcarpetas de fecha para evitar re-descargar archivos que ya existen."""
    existing: set[str] = set()
    pattern = re.compile(r"\d{8,}-\d+")
    if not base_download_dir.exists():
        return existing

    for date_dir in base_download_dir.iterdir():
        if not date_dir.is_dir():
            continue
        for candidate in date_dir.iterdir():
            if not candidate.is_dir() or not pattern.fullmatch(candidate.name):
                continue
            try:
                has_files = any(p.is_file() for p in candidate.iterdir())
            except Exception:
                has_files = False
            if has_files:
                existing.add(candidate.name)

    return existing


def _resolve_sqlite_path(base_download_dir: Path) -> Path:
    """Devuelve la ruta de SQLite (env: OSI_SQLITE_PATH) o fallback en downloads."""
    raw = (os.getenv("OSI_SQLITE_PATH") or "").strip()
    if raw:
        return Path(raw).expanduser().resolve()
    return (base_download_dir / "notificaciones.db").resolve()


def _resolve_default_download_dir() -> str:
    """Resuelve carpeta de descargas por defecto priorizando la carpeta sincronizada de SharePoint."""
    env_value = (os.getenv("OSI_DOWNLOAD_DIR") or "").strip()
    if env_value:
        return env_value

    user_profile = Path(os.getenv("USERPROFILE", "")).expanduser()
    if str(user_profile).strip():
        # Prioridad 1: carpeta sincronizada directamente con SharePoint via Teams
        sp_sync = user_profile / "Fonafe" / "CALIDAD-FISCALIZACION-ENSA - Notificaciones Osinergmin"
        if sp_sync.exists():
            return str(sp_sync)
        # Prioridad 2: carpeta OneDrive/OsiDoc (ruta anterior)
        preferred = user_profile / "OneDrive - Fonafe" / "OsiDoc" / "Documento"
        if preferred.exists():
            return str(preferred)

    return "downloads"


def _cleanup_empty_legacy_downloads(download_dir: Path) -> None:
    """Elimina carpeta local ./downloads si quedo vacia tras migracion."""
    workspace_dir = Path(__file__).resolve().parent
    legacy_dir = (workspace_dir / "downloads").resolve()
    target_dir = download_dir.resolve()

    if legacy_dir == target_dir or not legacy_dir.exists() or not legacy_dir.is_dir():
        return

    for folder in sorted([p for p in legacy_dir.rglob("*") if p.is_dir()], key=lambda p: len(p.parts), reverse=True):
        try:
            folder.rmdir()
        except Exception:
            continue

    try:
        legacy_dir.rmdir()
    except Exception:
        pass


def _migrate_legacy_local_downloads(download_dir: Path) -> int:
    """Mueve archivos rezagados de ./downloads y de la ruta anterior de OneDrive hacia la ruta configurada."""
    workspace_dir = Path(__file__).resolve().parent
    target_dir = download_dir.resolve()

    legacy_sources: list[Path] = []

    # Fuente 1: ./downloads local del proyecto
    legacy_local = (workspace_dir / "downloads").resolve()
    if legacy_local != target_dir and legacy_local.exists() and legacy_local.is_dir():
        legacy_sources.append(legacy_local)

    # Fuente 2: ruta anterior OneDrive/OsiDoc/Documento
    user_profile = Path(os.getenv("USERPROFILE", "")).expanduser()
    if str(user_profile).strip():
        old_onedrive = (user_profile / "OneDrive - Fonafe" / "OsiDoc" / "Documento").resolve()
        if old_onedrive != target_dir and old_onedrive.exists() and old_onedrive.is_dir():
            legacy_sources.append(old_onedrive)

    moved = 0
    for legacy_dir in legacy_sources:
        for source in legacy_dir.rglob("*"):
            if not source.is_file():
                continue
            if source.name.lower() == "notificaciones.db":
                continue
            # Ignorar archivos de metadata de SharePoint/OneDrive
            if source.name.startswith(".") and len(source.name) > 30:
                continue
            try:
                rel = source.relative_to(legacy_dir)
                target = target_dir / rel
                _move_file_with_retries(source, target)
                moved += 1
            except Exception:
                continue

        _cleanup_empty_legacy_downloads(legacy_dir)

    return moved


def _move_file_with_retries(source: Path, target: Path, retries: int = 10, delay_seconds: float = 0.35) -> Path:
    """Mueve archivo con reintentos para tolerar bloqueos transitorios de OneDrive/antivirus."""
    target.parent.mkdir(parents=True, exist_ok=True)
    for _ in range(max(1, retries)):
        try:
            if target.exists():
                target.unlink()
            source.replace(target)
            return target
        except Exception:
            time.sleep(delay_seconds)
    return source


def _normalize_exported_excel_file(download_dir: Path, downloaded_file: Path, processing_date: str) -> Path:
    """Normaliza exportaciones Excel para que siempre queden dentro de la carpeta por fecha."""
    day_dir = download_dir / processing_date
    day_dir.mkdir(parents=True, exist_ok=True)
    canonical = day_dir / f"Notificaciones Electrónicas - {processing_date}.xlsx"

    # 1) El archivo detectado en este ciclo debe quedar con nombre diario canonico.
    try:
        if downloaded_file.exists():
            downloaded_file = _move_file_with_retries(downloaded_file, canonical)
    except Exception:
        downloaded_file = canonical if canonical.exists() else downloaded_file

    # 2) Limpieza defensiva: mueve cualquier Excel suelto en la raiz de downloads.
    dated_pattern = re.compile(r"^Notificaciones Electr[oó]nicas - (\d{4}-\d{2}-\d{2})\.xlsx$", re.IGNORECASE)
    generic_pattern = re.compile(r"^Notificaciones Electr[oó]nicas(?: \(\d+\))?\.xlsx$", re.IGNORECASE)
    for candidate in download_dir.glob("*.xlsx"):
        name = candidate.name
        match = dated_pattern.fullmatch(name)
        if match:
            token = match.group(1)
            target = download_dir / token / f"Notificaciones Electrónicas - {token}.xlsx"
            _move_file_with_retries(candidate, target)
            continue

        if generic_pattern.fullmatch(name):
            _move_file_with_retries(candidate, canonical)

    dup_pattern = re.compile(r"^Notificaciones Electr[oó]nicas \(\d+\)\.xlsx$", re.IGNORECASE)
    for candidate in download_dir.glob("*.xlsx"):
        name = candidate.name
        if not dup_pattern.fullmatch(name):
            continue
        try:
            candidate.unlink()
        except Exception:
            pass

    return canonical if canonical.exists() else downloaded_file


def _enforce_daily_excel_layout(download_dir: Path, processing_date: str) -> Path:
    """Garantiza que el Excel diario quede en carpeta YYYY-MM-DD con nombre canonico."""
    day_dir = download_dir / processing_date
    day_dir.mkdir(parents=True, exist_ok=True)
    canonical = day_dir / f"Notificaciones Electrónicas - {processing_date}.xlsx"

    if canonical.exists():
        return canonical

    generic_pattern = re.compile(r"^Notificaciones Electr[oó]nicas(?: \(\d+\))?\.xlsx$", re.IGNORECASE)
    dated_pattern = re.compile(r"^Notificaciones Electr[oó]nicas - (\d{4}-\d{2}-\d{2})\.xlsx$", re.IGNORECASE)

    for candidate in download_dir.glob("*.xlsx"):
        if not candidate.is_file():
            continue

        name = candidate.name
        dated_match = dated_pattern.fullmatch(name)
        if dated_match:
            token = dated_match.group(1)
            target = download_dir / token / f"Notificaciones Electrónicas - {token}.xlsx"
            _move_file_with_retries(candidate, target)
            continue

        if generic_pattern.fullmatch(name):
            _move_file_with_retries(candidate, canonical)

    return canonical


def _relocate_stray_root_files(download_dir: Path, processing_date: str) -> int:
    """Mueve cualquier archivo suelto en la raiz de download_dir a la subcarpeta de fecha.
    Evita que archivos descargados sin mover aparezcan en la raiz de SharePoint al sincronizar."""
    if not download_dir.exists():
        return 0

    excel_pattern = re.compile(r"^Notificaciones Electr[oó]nicas.*\.xlsx$", re.IGNORECASE)
    dest_dir = download_dir / processing_date / "sin-clasificar"
    moved = 0

    for candidate in list(download_dir.iterdir()):
        if candidate.is_dir():
            continue
        # Los .db se quedan donde estan (gestionados por OSI_SQLITE_PATH)
        if candidate.suffix.lower() == ".db":
            continue
        # Los Excel se manejan en _enforce_daily_excel_layout
        if excel_pattern.match(candidate.name):
            continue

        dest_dir.mkdir(parents=True, exist_ok=True)
        target = dest_dir / candidate.name
        try:
            _move_file_with_retries(candidate, target)
            logging.info("Archivo suelto movido a carpeta de fecha: %s → %s", candidate.name, target.relative_to(download_dir))
            moved += 1
        except Exception:
            continue

    return moved


def _create_empty_daily_excel(download_dir: Path, processing_date: str) -> Path:
    """Crea un Excel diario vacio para marcar que no hubo datos ese dia."""
    day_dir = download_dir / processing_date
    day_dir.mkdir(parents=True, exist_ok=True)

    if Workbook is None:
        output_path = day_dir / f"Notificaciones Electrónicas - {processing_date}.csv"
        if not output_path.exists():
            output_path.write_text("", encoding="utf-8")
        return output_path

    output_path = day_dir / f"Notificaciones Electrónicas - {processing_date}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Notificaciones"
    wb.save(output_path)
    return output_path


def _update_excel_control_sheet(
    excel_path: Path,
    processing_date: str,
    notifications_count: int,
    docs_downloaded: int,
) -> None:
    """Actualiza una hoja de control para dejar trazabilidad del ultimo proceso."""
    if Workbook is None:
        return

    if excel_path.suffix.lower() != ".xlsx" or not excel_path.exists():
        return

    try:
        import openpyxl  # type: ignore[import-not-found]

        wb = openpyxl.load_workbook(excel_path)
        if "Control" in wb.sheetnames:
            ws = wb["Control"]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet("Control")

        ws.append(["processing_date", "updated_at", "notificaciones", "documentos_descargados"])
        ws.append([
            processing_date,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            int(max(0, notifications_count)),
            int(max(0, docs_downloaded)),
        ])

        wb.save(excel_path)
        wb.close()
    except Exception as exc:
        logging.warning("No se pudo actualizar la hoja Control en %s: %s", excel_path.name, exc)


def _cleanup_day_notification_folders(base_download_dir: Path, processing_date: str) -> int:
    """Elimina carpetas de notificaciones del dia cuando la consulta no devuelve datos."""
    day_dir = base_download_dir / processing_date
    if not day_dir.exists() or not day_dir.is_dir():
        return 0

    removed = 0
    for candidate in day_dir.iterdir():
        if not candidate.is_dir():
            continue
        try:
            shutil.rmtree(candidate)
            removed += 1
        except Exception:
            continue
    return removed


def _clear_processing_date_rows(db_path: Path, processing_date: str) -> None:
    """Limpia filas del dia en SQLite cuando no hubo resultados en esa fecha."""
    if not processing_date:
        return

    if not db_path.exists():
        return

    try:
        con = sqlite3.connect(str(db_path))
        cur = con.cursor()
        existing_cols = {row[1].lower() for row in cur.execute('PRAGMA table_info(notificaciones)')}
        if "processing_date" in existing_cols:
            cur.execute('DELETE FROM notificaciones WHERE processing_date = ?', (processing_date,))
            con.commit()
        con.close()
    except Exception as exc:
        logging.warning("No se pudo limpiar SQLite para fecha %s: %s", processing_date, exc)


def _save_excel_to_sqlite(excel_path: Path, db_path: Path, processing_date: str | None = None) -> int:
    """Lee el Excel exportado y guarda/actualiza filas en SQLite separadas por dia."""
    try:
        import openpyxl  # type: ignore[import-not-found]
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active

        rows_iter = ws.iter_rows(values_only=True)
        raw_headers = next(rows_iter, None)
        if raw_headers is None:
            wb.close()
            logging.warning("Excel vacio, no se guardo nada en SQLite.")
            return 0

        # Normaliza cabeceras para usar como columnas SQL.
        def _col_name(h: object, idx: int) -> str:
            s = re.sub(r"[^\w]", "_", unicodedata.normalize("NFD", str(h or f"col_{idx}"))
                       .encode("ascii", "ignore").decode())
            return (s.strip("_") or f"col_{idx}").lower()

        col_names = [_col_name(h, i) for i, h in enumerate(raw_headers)]
        seen_cols: set[str] = set()
        final_cols: list[str] = []
        for c in col_names:
            base = c or "col"
            name = base
            n = 1
            while name in seen_cols:
                name = f"{base}_{n}"
                n += 1
            seen_cols.add(name)
            final_cols.append(name)

        db_path.parent.mkdir(parents=True, exist_ok=True)
        con = sqlite3.connect(str(db_path))
        cur = con.cursor()

        col_defs = ", ".join(f'"{c}" TEXT' for c in final_cols)
        cur.execute(f'CREATE TABLE IF NOT EXISTS notificaciones ({col_defs})')

        existing_cols = {row[1].lower() for row in cur.execute('PRAGMA table_info(notificaciones)')}
        for c in final_cols:
            if c.lower() not in existing_cols:
                cur.execute(f'ALTER TABLE notificaciones ADD COLUMN "{c}" TEXT')

        if "fecha_importacion" not in existing_cols:
            try:
                cur.execute('ALTER TABLE notificaciones ADD COLUMN "fecha_importacion" TEXT')
            except sqlite3.OperationalError:
                pass

        if "processing_date" not in existing_cols:
            try:
                cur.execute('ALTER TABLE notificaciones ADD COLUMN "processing_date" TEXT')
            except sqlite3.OperationalError:
                pass

        # Encuentra indice de la columna de fecha de notificacion para asignar processing_date real.
        notif_date_col_idx: int | None = None
        for _i, _c in enumerate(final_cols):
            if _c.lower() in {"fecha_de_notificacion", "fecha_notificacion"}:
                notif_date_col_idx = _i
                break

        _date_fmts = (
            "%d/%m/%Y %I:%M:%S %p",
            "%d/%m/%Y %H:%M:%S",
            "%d/%m/%Y",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d",
        )

        def _derive_row_date(values: list[str]) -> str:
            if notif_date_col_idx is not None and notif_date_col_idx < len(values):
                raw = values[notif_date_col_idx].strip()
                for _fmt in _date_fmts:
                    try:
                        return datetime.strptime(raw, _fmt).strftime("%Y-%m-%d")
                    except (ValueError, AttributeError):
                        continue
            return processing_date or ""

        # Lee todas las filas en memoria para determinar las fechas reales antes de limpiar BD.
        all_data_rows: list[list[str]] = []
        for data_row in rows_iter:
            vals = [str(v).strip() if v is not None else "" for v in data_row[:len(final_cols)]]
            while len(vals) < len(final_cols):
                vals.append("")
            all_data_rows.append(vals)

        row_dates = [_derive_row_date(v) for v in all_data_rows]

        # Limpia registros de la fecha filtro Y de las fechas reales encontradas en los datos.
        dates_to_clean: set[str] = {d for d in row_dates if d}
        if processing_date:
            dates_to_clean.add(processing_date)
        for _d in dates_to_clean:
            try:
                cur.execute('DELETE FROM notificaciones WHERE processing_date = ?', (_d,))
            except sqlite3.OperationalError:
                pass

        placeholders = ", ".join("?" for _ in final_cols)
        col_list = ", ".join(f'"{c}"' for c in final_cols)
        insert_sql = (
            f'INSERT INTO notificaciones ({col_list}, "fecha_importacion", "processing_date") '
            f'VALUES ({placeholders}, ?, ?)'
        )

        fecha_importacion = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        count = 0
        for values, row_date in zip(all_data_rows, row_dates):
            cur.execute(insert_sql, values + [fecha_importacion, row_date])
            count += 1

        con.commit()
        con.close()
        wb.close()
        logging.info(
            "SQLite actualizado (%s): %s filas guardadas en tabla 'notificaciones' para %s.",
            db_path.name,
            count,
            processing_date or "sin-fecha",
        )
        return count
    except Exception as exc:
        logging.warning("Error guardando Excel en SQLite: %s", exc)
        return 0


def _get_notification_numbers_from_excel(excel_path: Path) -> list[str]:
    """Lee todos los Nro. Notificacion desde el Excel exportado (primera columna)."""
    try:
        import openpyxl  # type: ignore[import-not-found]
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        numbers: list[str] = []
        nro_col: int | None = None
        for row in ws.iter_rows(values_only=True):
            if nro_col is None:
                # Detecta la columna 'Nro. Notificacion' en la cabecera
                for i, cell in enumerate(row):
                    if cell and "notif" in str(cell).lower():
                        nro_col = i
                        break
                if nro_col is None:
                    # Sin cabecera reconocible, usa columna 0
                    nro_col = 0
                continue
            val = re.sub(r"[^0-9\-]", "", str(row[nro_col] or "")).strip("-")
            if re.fullmatch(r"\d{8,}-\d+", val):
                numbers.append(val)
        wb.close()
        # Deduplica preservando orden
        seen: set[str] = set()
        unique_numbers: list[str] = []
        for n in numbers:
            if n not in seen:
                seen.add(n)
                unique_numbers.append(n)
        logging.info("Numeros de notificacion leidos desde Excel (%s): %s unicos de %s totales", excel_path.name, len(unique_numbers), len(numbers))
        return unique_numbers
    except Exception as exc:
        logging.warning("No se pudo leer numeros de notificacion desde Excel '%s': %s", excel_path, exc)
        return []


def _get_visible_notification_numbers(driver) -> list[str]:
    """Obtiene numeros de notificacion visibles unicos del grid actual."""
    try:
        raw_list = driver.execute_script(
            """
const isVisible = (el) => {
    if (!el) return false;
    const style = window.getComputedStyle(el);
    return style.display !== 'none' && style.visibility !== 'hidden' && el.offsetParent !== null;
};
const rows = Array.from(document.querySelectorAll('table.ui-jqgrid-btable tr.jqgrow')).filter((tr) => isVisible(tr));
const out = [];
for (const row of rows) {
    const cells = Array.from(row.querySelectorAll('td')).map((td) => (td.textContent || '').trim());
    const direct = cells.find((v) => /\\d{8,}-\\d+/.test(v));
    if (direct) {
        out.push(direct);
        continue;
    }
    const rowText = (row.textContent || '').replace(/\\s+/g, ' ').trim();
    const m = rowText.match(/\\d{8,}-\\d+/);
    if (m) out.push(m[0]);
}
return out;
"""
        )
        numbers = []
        for item in raw_list or []:
            clean = re.sub(r"[^0-9\-]", "", str(item or "")).strip("-")
            if clean:
                numbers.append(clean)
        # Unicos preservando orden
        seen: set[str] = set()
        unique = []
        for n in numbers:
            if n in seen:
                continue
            seen.add(n)
            unique.append(n)
        return unique
    except Exception:
        return []


def _download_visible_document_links_for_notification(
    driver,
    cfg: AuthConfig,
    download_dir: Path,
    notification_number: str,
    processing_date: str,
) -> int:
    """Descarga links visibles de documentos y los organiza por numero de notificacion."""
    downloaded = 0

    def _get_unique_links() -> list[dict[str, str]]:
        try:
            raw_links = driver.execute_script(
                """
const isVisible = (el) => {
    if (!el) return false;
    const style = window.getComputedStyle(el);
    return style.display !== 'none' && style.visibility !== 'hidden' && el.offsetParent !== null;
};
const links = Array.from(document.querySelectorAll("a[title='Descargar archivo'], a[href*='descargarArchivoNotificacion']"))
    .filter((a) => isVisible(a));
const out = [];
const seen = new Set();
for (const a of links) {
    const href = (a.getAttribute('href') || '').trim();
    const text = (a.textContent || '').trim();
    const key = href || text;
    if (!key || seen.has(key)) continue;
    seen.add(key);
    out.push({ href, text });
}
return out;
"""
            )
            return list(raw_links or [])
        except Exception:
            return []

    unique_links = _get_unique_links()
    for link_data in unique_links:
        href = str((link_data or {}).get("href", "") or "")
        text = str((link_data or {}).get("text", "") or "")
        if not href and not text:
            continue

        before = _snapshot_downloads(download_dir)
        try:
            clicked = bool(
                driver.execute_script(
                    """
const wantedHref = arguments[0];
const wantedText = arguments[1];
const isVisible = (el) => {
    if (!el) return false;
    const style = window.getComputedStyle(el);
    return style.display !== 'none' && style.visibility !== 'hidden' && el.offsetParent !== null;
};
const links = Array.from(document.querySelectorAll("a[title='Descargar archivo'], a[href*='descargarArchivoNotificacion']"))
    .filter((a) => isVisible(a));
const link = links.find((a) => {
    const href = (a.getAttribute('href') || '').trim();
    const text = (a.textContent || '').trim();
    return (wantedHref && href === wantedHref) || (!wantedHref && wantedText && text === wantedText);
});
if (!link) return false;
try { link.scrollIntoView({ block: 'center' }); } catch (e) {}
try { link.click(); } catch (e) {}
return true;
""",
                    href,
                    text,
                )
            )
        except Exception:
            clicked = False

        if not clicked:
            continue

        _accept_browser_alert_if_present(driver)
        # Delay para que el servidor inicie la descarga
        time.sleep(1.5)
        # Timeout más largo para PDFs
        file_path = _wait_for_new_download(download_dir, before, max(60, cfg.export_wait_seconds))
        if file_path is not None:
            # Verifica que el archivo tenga contenido
            try:
                file_size = file_path.stat().st_size
                if file_size < 100:
                    logging.warning("Archivo descargado muy pequeño (%d bytes), probablemente corrupto: %s", file_size, file_path.name)
                    continue
            except OSError as e:
                logging.warning("No se pudo verificar tamaño del archivo %s: %s", file_path.name, e)
                continue
            
            final_path = _move_download_to_notification_folder(
                file_path,
                download_dir,
                notification_number,
                processing_date,
            )
            logging.info(
                "Archivo movido a carpeta de suministro %s en fecha %s: %s (%d bytes)",
                notification_number or "sin-numero",
                processing_date,
                final_path.name,
                file_size,
            )
            downloaded += 1

    return downloaded


def _close_visible_dialogs(driver) -> None:
    """Intenta cerrar dialogs/modales visibles para volver al listado."""
    try:
        driver.execute_script(
            """
const isVisible = (el) => {
    if (!el) return false;
    const style = window.getComputedStyle(el);
    return style.display !== 'none' && style.visibility !== 'hidden' && el.offsetParent !== null;
};
const closeCandidates = Array.from(document.querySelectorAll(
  "a.ui-dialog-titlebar-close, button.ui-dialog-titlebar-close, .ui-icon-closethick, button, a"
)).filter((el) => isVisible(el));

for (const el of closeCandidates) {
  const text = (el.innerText || el.textContent || '').trim().toLowerCase();
  if (el.classList.contains('ui-dialog-titlebar-close') || text === 'cerrar' || text === 'volver' || text === 'regresar') {
    try { el.click(); } catch (e) {}
  }
}
"""
        )
    except Exception:
        pass


def _click_regresar_sequence(driver, cfg: AuthConfig) -> None:
    """Hace clic en Regresar de documentos y luego en Regresar de detalle."""

    def _click_if_present(selectors: list[tuple[str, str]]) -> bool:
        target = _find_first(driver, selectors, 3)
        if target is None:
            return False

        try:
            if _click_ingresar_button(driver, target):
                return True
        except Exception:
            pass

        try:
            driver.execute_script("arguments[0].click();", target)
            return True
        except Exception:
            return False

    # 1) Regresar de Documentos notificados
    _click_if_present(
        [
            (By.ID, "regresar-documentosNotificacion-boton"),
            (By.XPATH, "//input[@type='button' and @id='regresar-documentosNotificacion-boton' and @value='Regresar']"),
        ]
    )
    time.sleep(0.35)

    # 2) Regresar del detalle de notificacion
    _click_if_present(
        [
            (By.ID, "regresar-boton"),
            (By.XPATH, "//input[@type='button' and @id='regresar-boton' and @value='Regresar']"),
        ]
    )
    time.sleep(0.35)


def _download_documents_from_visible_results(
    driver,
    cfg: AuthConfig,
    download_dir: Path,
    processing_date: str,
    notification_numbers: list[str] | None = None,
) -> int:
    """Por cada notificacion (del Excel o del grid), abre detalle y descarga sus documentos."""
    total_downloads = 0
    if notification_numbers is None:
        notification_numbers = _get_visible_notification_numbers(driver)
        logging.info("Notificaciones leidas del grid visible: %s", len(notification_numbers))
    else:
        logging.info("Notificaciones leidas del Excel exportado: %s", len(notification_numbers))
    if not notification_numbers:
        return 0

    for idx, notification_number in enumerate(notification_numbers, start=1):
        logging.info("Procesando notificacion %s de %s: %s", idx, len(notification_numbers), notification_number)

        downloaded_here = 0
        max_open_attempts = 5
        opened_docs = False
        for attempt in range(1, max_open_attempts + 1):
            if not _click_lupa_by_notification_number_any_page(driver, cfg, notification_number):
                if attempt == max_open_attempts:
                    logging.warning("No se pudo abrir la lupita de la notificacion %s.", notification_number)
                    break
                logging.info(
                    "Reintento %s/%s para abrir lupita de notificacion %s.",
                    attempt + 1,
                    max_open_attempts,
                    notification_number,
                )
                _close_visible_dialogs(driver)
                time.sleep(0.4)
                continue

            if not _click_documentos_notificados(driver, cfg):
                if attempt == max_open_attempts:
                    logging.warning("No se pudo abrir 'Documentos notificados' en la notificacion %s.", notification_number)
                    _click_regresar_sequence(driver, cfg)
                    _close_visible_dialogs(driver)
                    break
                logging.info(
                    "Reintento %s/%s para abrir 'Documentos notificados' en %s.",
                    attempt + 1,
                    max_open_attempts,
                    notification_number,
                )
                _click_regresar_sequence(driver, cfg)
                _close_visible_dialogs(driver)
                time.sleep(0.4)
                continue

            time.sleep(0.4)
            downloaded_here = _download_visible_document_links_for_notification(
                driver,
                cfg,
                download_dir,
                notification_number,
                processing_date,
            )
            opened_docs = True
            break

        total_downloads += downloaded_here
        logging.info("Documentos descargados en notificacion %s: %s", notification_number, downloaded_here)
        if opened_docs:
            _click_regresar_sequence(driver, cfg)
            _close_visible_dialogs(driver)
            time.sleep(0.4)

    return total_downloads


def _get_visible_grid_status(driver) -> tuple[int, str, bool, bool]:
    """Lee estado visible del listado: filas, pager, bandera sin resultados y bandera pager con datos."""
    try:
        status = driver.execute_script(
            """
const normalize = (v) => (v || '')
    .normalize('NFD')
    .replace(/[\u0300-\u036f]/g, '')
    .replace(/\u00a0/g, ' ')
    .replace(/\\s+/g, ' ')
    .trim()
    .toLowerCase();
const clean = (v) => (v || '').replace(/\u00a0/g, ' ').replace(/\\s+/g, ' ').trim();
const isVisible = (el) => {
    if (!el) return false;
    const style = window.getComputedStyle(el);
    return style.display !== 'none' && style.visibility !== 'hidden' && el.offsetParent !== null;
};

const rows = Array.from(document.querySelectorAll('table.ui-jqgrid-btable tr.jqgrow'))
    .filter((tr) => isVisible(tr))
    .map((tr) => Array.from(tr.querySelectorAll('td')).map((td) => clean(td.textContent)));

const dataRows = rows.filter((cells) => {
    // Ignora celdas vacias (checkbox/acciones) y exige al menos una celda con contenido real.
    const realCells = cells.slice(2);
    return realCells.some((v) => v);
});

const pagerText = (Array.from(document.querySelectorAll("td[id$='_right'], .ui-paging-info"))
    .map((el) => clean(el.textContent))
    .find(Boolean)) || '';

const noResultText = (Array.from(document.querySelectorAll("td[id$='_records'], td[id$='_right'], .ui-paging-info, .ui-jqgrid-bdiv"))
    .map((el) => normalize(el.textContent))
    .find((t) => t.includes('no se encontro resultados') || t.includes('no se encontraron resultados'))) || '';

const hasPagerResults = /^mostrando\\s+\\d+\\s*-\\s*\\d+\\s+de\\s+\\d+$/i.test(clean(pagerText));

return {
    rowCount: dataRows.length,
    pagerText,
    noResults: !!noResultText,
    hasPagerResults,
};
"""
        )
        return (
            int((status or {}).get("rowCount", 0)),
            str((status or {}).get("pagerText", "")),
            bool((status or {}).get("noResults", False)),
            bool((status or {}).get("hasPagerResults", False)),
        )
    except Exception:
        return 0, "", False, False


def _export_filtered_rows_to_excel(download_dir: Path, headers: list[str], rows: list[list[str]]) -> Path:
    """Genera un Excel local con el resultado exacto del grid filtrado."""
    download_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    if Workbook is None:
        output_path = download_dir / f"Notificaciones Filtradas {stamp}.csv"
        with output_path.open("w", newline="", encoding="utf-8-sig") as fh:
            writer = csv.writer(fh)
            if headers:
                writer.writerow(headers)
            writer.writerows(rows)
        return output_path

    output_path = download_dir / f"Notificaciones Filtradas {stamp}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Notificaciones Filtradas"

    if headers:
        ws.append(headers)
    for row in rows:
        ws.append(row)

    wb.save(output_path)
    return output_path


def _apply_sne_filters(driver, cfg: AuthConfig) -> bool:
    """Llena fechas y ejecuta la busqueda en la bandeja del SNE."""
    now = datetime.now()
    today = now.strftime("%d/%m/%Y")
    # Por defecto se procesa solo el dia actual para evitar arrastrar historico.
    default_inicio = today
    fecha_inicio = cfg.fecha_notificacion_inicio or default_inicio
    fecha_fin = cfg.fecha_notificacion_fin or today
    processing_date = _resolve_processing_date(cfg)

    if not fecha_inicio or not fecha_fin:
        logging.warning("Se omite la busqueda en SNE porque falta una de las dos fechas requeridas.")
        return False

    inicio = _find_first(driver, [(By.ID, cfg.sne_fecha_inicio_id)], cfg.timeout)
    fin = _find_first(driver, [(By.ID, cfg.sne_fecha_fin_id)], cfg.timeout)
    buscar = _find_first(
        driver,
        [
            (By.ID, cfg.sne_buscar_button_id),
            (By.XPATH, "//input[@type='button' and @id='buscar-boton']"),
            (By.XPATH, "//input[@type='button' and @value='Buscar']"),
        ],
        cfg.timeout,
    )

    if inicio is None or fin is None or buscar is None:
        logging.warning("No se encontraron los campos de fecha o el boton Buscar en la bandeja del SNE.")
        return False

    def _apply_filter_values() -> None:
        # Usa la API de jQuery datepicker para asegurar la fecha correcta (evita interferencia del picker).
        for campo, valor in [(inicio, fecha_inicio), (fin, fecha_fin)]:
            try:
                driver.execute_script(
                    """
const el = arguments[0];
const val = arguments[1];
el.value = val;
if (window.jQuery) {
    try {
        const parts = val.split('/');
        const d = new Date(parseInt(parts[2]), parseInt(parts[1]) - 1, parseInt(parts[0]));
        jQuery(el).datepicker('setDate', d);
    } catch(e) {}
}
['input', 'change', 'blur'].forEach((t) => {
    el.dispatchEvent(new Event(t, { bubbles: true }));
});
""",
                    campo,
                    valor,
                )
            except Exception:
                pass

        # Selecciona estado de lectura en el filtro del SNE.
        try:
            select_leido = driver.find_element(By.ID, "leidoNotificacion")
            # Fuerza siempre "Todos" para mantener el comportamiento solicitado.
            leido_value = ""
            driver.execute_script(
                """
const sel = arguments[0];
const val = arguments[1];
sel.value = val;
['change', 'input'].forEach((t) => sel.dispatchEvent(new Event(t, { bubbles: true })));
""",
                select_leido,
                leido_value,
            )
            logging.info("Filtro de lectura aplicado en leidoNotificacion: '%s'.", leido_value or "Todos")
        except Exception as exc:
            logging.warning("No se pudo seleccionar estado de lectura en leidoNotificacion: %s", exc)

    logging.info(
        "Fechas de notificacion cargadas en SNE: inicio=%s fin=%s",
        fecha_inicio,
        fecha_fin,
    )

    headers: list[str] = []
    rows: list[list[str]] = []
    max_attempts = 4
    for intento in range(1, max_attempts + 1):
        _apply_filter_values()

        # Replica el comportamiento manual: dos clics secuenciales en Buscar.
        search_clicked = False
        visible_rows = 0
        pager_text = ""
        no_results = False
        has_pager_results = False
        for subintento in range(1, 4):
            first_click = _click_search_button_and_wait(driver, buscar, cfg)
            time.sleep(1.2)
            second_click = _click_search_button_and_wait(driver, buscar, cfg)

            search_clicked = search_clicked or first_click or second_click

            # Espera corta de estabilizacion para que jqGrid renderice filas visibles.
            end_wait = time.time() + max(2.5, cfg.timeout / 3)
            while time.time() < end_wait:
                visible_rows, pager_text, no_results, has_pager_results = _get_visible_grid_status(driver)
                if visible_rows > 0 and not no_results and has_pager_results:
                    break
                time.sleep(0.25)

            logging.info(
                "Estado visible del grid tras doble Buscar (ciclo %s/3, intento %s/%s): filas=%s, pager='%s', no_results=%s, pager_ok=%s",
                subintento,
                intento,
                max_attempts,
                visible_rows,
                pager_text,
                no_results,
                has_pager_results,
            )

            if visible_rows > 0 and not no_results and has_pager_results:
                break

            time.sleep(0.8)

        if not search_clicked:
            if intento == max_attempts:
                logging.warning("No se pudo hacer clic en el boton Buscar del SNE.")
                return False
            time.sleep(0.5)
            continue

        if visible_rows <= 0 or no_results or not has_pager_results:
            if intento < max_attempts:
                logging.warning("Grid visible sin datos en intento %s; reintentando ciclo completo.", intento)
                time.sleep(1.0)
                continue

        headers, rows = _collect_filtered_grid_rows(driver, cfg)
        if rows:
            break

        if intento < max_attempts:
            logging.warning("Busqueda sin filas en intento %s; reintentando clic en Buscar.", intento)
            time.sleep(0.8)

    download_dir = Path(cfg.download_dir).expanduser().resolve()
    download_dir.mkdir(parents=True, exist_ok=True)
    migrated = _migrate_legacy_local_downloads(download_dir)
    if migrated > 0:
        logging.info("Migracion automatica aplicada: %s archivo(s) movido(s) desde ./downloads.", migrated)
    _enforce_daily_excel_layout(download_dir, processing_date)
    stray = _relocate_stray_root_files(download_dir, processing_date)
    if stray > 0:
        logging.info("Archivos sueltos en raiz movidos a carpeta de fecha: %s archivo(s).", stray)

    if not rows:
        empty_file = _create_empty_daily_excel(download_dir, processing_date)
        _update_excel_control_sheet(empty_file, processing_date, notifications_count=0, docs_downloaded=0)
        removed = _cleanup_day_notification_folders(download_dir, processing_date)
        _clear_processing_date_rows(_resolve_sqlite_path(download_dir), processing_date)
        logging.info(
            "Sin resultados en la busqueda. Archivo diario vacio: %s. Carpetas del dia limpiadas: %s",
            empty_file.name,
            removed,
        )
        return True

    downloads_before = _snapshot_downloads(download_dir)

    export_btn = _find_first(
        driver,
        [
            (By.XPATH, "//div[@class='ui-pg-div' and .//span[contains(@class,'ui-icon-arrowthickstop-1-s')] and contains(normalize-space(.), 'Exportar a Excel')]"),
            (By.XPATH, cfg.sne_export_excel_selector),
            (By.XPATH, "//div[contains(@class,'ui-pg-div')][.//span[contains(@class,'ui-icon-arrowthickstop-1-s')]]"),
            (By.XPATH, "//div[contains(@class,'ui-pg-div') and contains(normalize-space(.), 'Exportar a Excel')]"),
        ],
        cfg.timeout,
    )
    if export_btn is None:
        logging.warning("No se encontro el boton Exportar a Excel en la bandeja del SNE.")
        return False

    if not _click_export_excel_button(driver, export_btn):
        logging.warning("No se pudo hacer clic en Exportar a Excel.")
        return False

    time.sleep(0.25)
    alert_text = _accept_browser_alert_if_present(driver, expected_text="No hay datos para exportar")
    if "no hay datos para exportar" in (alert_text or "").lower():
        empty_file = _create_empty_daily_excel(download_dir, processing_date)
        _update_excel_control_sheet(empty_file, processing_date, notifications_count=0, docs_downloaded=0)
        removed = _cleanup_day_notification_folders(download_dir, processing_date)
        _clear_processing_date_rows(_resolve_sqlite_path(download_dir), processing_date)
        logging.info(
            "La plataforma reporto que no hay datos para exportar. Archivo diario vacio: %s. Carpetas del dia limpiadas: %s",
            empty_file.name,
            removed,
        )
        return True

    downloaded_file = _wait_for_new_download(download_dir, downloads_before, cfg.export_wait_seconds)
    if downloaded_file is None:
        logging.warning("No se detecto descarga de Excel dentro del tiempo esperado (%ss).", cfg.export_wait_seconds)
        return False

    downloaded_file = _normalize_exported_excel_file(download_dir, downloaded_file, processing_date)

    logging.info("Exportacion a Excel completada. Archivo descargado: %s", downloaded_file.name)

    # Guarda todas las filas del Excel en SQLite (aislado por fecha de procesamiento).
    db_path = _resolve_sqlite_path(download_dir)
    _save_excel_to_sqlite(downloaded_file, db_path, processing_date)

    # Usa el Excel como fuente de verdad: contiene TODOS los Nro. Notificacion
    # independientemente de la paginacion del grid.
    excel_notification_numbers = _get_notification_numbers_from_excel(downloaded_file)
    if not excel_notification_numbers:
        logging.warning("No se leyeron numeros de notificacion desde el Excel; se usara el grid visible.")
        excel_notification_numbers = None

    if cfg.target_notifications:
        if excel_notification_numbers is None:
            excel_notification_numbers = list(cfg.target_notifications)
        else:
            wanted = set(cfg.target_notifications)
            excel_notification_numbers = [n for n in excel_notification_numbers if n in wanted]
        logging.info(
            "Modo focalizado activo: %s notificaciones objetivo.",
            len(excel_notification_numbers or []),
        )

    if excel_notification_numbers and (cfg.incremental_only or cfg.skip_existing_notifications):
        already_downloaded = _get_notifications_with_downloads(download_dir, processing_date)
        total_before = len(excel_notification_numbers)
        excel_notification_numbers = [n for n in excel_notification_numbers if n not in already_downloaded]
        logging.info(
            "Modo incremental: %s pendientes de %s totales (ya descargadas: %s).",
            len(excel_notification_numbers),
            total_before,
            max(0, total_before - len(excel_notification_numbers)),
        )
        if not excel_notification_numbers:
            logging.info("No hay notificaciones nuevas o pendientes por descargar.")
            return True

    docs_downloaded = _download_documents_from_visible_results(
        driver,
        cfg,
        download_dir,
        processing_date,
        excel_notification_numbers,
    )
    if docs_downloaded > 0:
        logging.info("Descarga de documentos notificados completada. Archivos descargados: %s", docs_downloaded)
    else:
        logging.warning("No se descargaron documentos notificados.")

    _update_excel_control_sheet(
        downloaded_file,
        processing_date,
        notifications_count=len(excel_notification_numbers or []),
        docs_downloaded=docs_downloaded,
    )

    return True


def _resolve_sne_menu_targets(driver, menu):
    if menu is None:
        return None, None

    try:
        container = driver.execute_script(
            """
const el = arguments[0];
if (!el) return null;
return el.closest('mat-list-item, .mat-mdc-list-item, .mdc-list-item, [role="listitem"]') || el;
""",
            menu,
        )
    except Exception:
        container = menu

    try:
        title = driver.execute_script(
            """
const el = arguments[0];
if (!el) return null;
return el.querySelector('.mat-mdc-list-item-title, .text-menu-parent') || el;
""",
            container,
        )
    except Exception:
        title = menu

    return container, title


def _find_sne_menu(driver, cfg: AuthConfig, per_selector_timeout: int = 2):
    menu = _find_first(
        driver,
        [
            (By.XPATH, cfg.sne_menu_selector),
            (By.CSS_SELECTOR, "div[matlistitemtitle].text-menu-parent"),
            (By.CSS_SELECTOR, "div[matlistitemtitle].text-menu"),
            (
                By.XPATH,
                "//div[@matlistitemtitle and contains(@class,'text-menu') "
                "and contains(normalize-space(),'SNE')]",
            ),
            (
                By.XPATH,
                "//div[@matlistitemtitle and contains(@class,'mat-mdc-list-item-title') "
                "and contains(normalize-space(),'Casilla') and contains(normalize-space(),'SNE')]",
            ),
            (
                By.XPATH,
                "//mat-list-item[.//mat-icon[@data-mat-icon-name='ico-mail'] "
                "and .//div[contains(normalize-space(),'SNE')]]",
            ),
            (
                By.XPATH,
                "//*[contains(normalize-space(.), 'Casilla') and contains(normalize-space(.), 'SNE') "
                "and (self::div or self::span or self::a)]",
            ),
        ],
        per_selector_timeout,
    )

    if menu is not None:
        return menu

    # Fallback robusto: ubica el titulo visible exacto y devuelve su mat-list-item padre.
    try:
        return driver.execute_script(
            """
const normalize = (v) => (v || '').normalize('NFD').replace(/[\u0300-\u036f]/g, '').trim().toLowerCase();
const wanted = normalize('Casilla Electrónica del SNE');
const nodes = Array.from(document.querySelectorAll('div[matlistitemtitle].text-menu-parent, div[matlistitemtitle].text-menu, div[matlistitemtitle].mat-mdc-list-item-title, .mat-mdc-list-item-title'));

for (const node of nodes) {
  const text = normalize(node.textContent || node.innerText || '');
  const style = window.getComputedStyle(node);
  const visible = style && style.visibility !== 'hidden' && style.display !== 'none' && node.offsetParent !== null;
  if (visible && text === wanted) {
    return node.closest('mat-list-item, .mat-mdc-list-item, .mdc-list-item') || node;
  }
}

return null;
"""
        )
    except Exception:
        return None


def _force_click_sne_menu_via_js(driver) -> bool:
    """Ultimo recurso: clic JS directo sobre el item visible de Casilla SNE."""
    try:
        clicked = driver.execute_script(
            """
const normalize = (v) => (v || '').normalize('NFD').replace(/[\u0300-\u036f]/g, '').trim().toLowerCase();
const wanted = normalize('Casilla Electrónica del SNE');
const nodes = Array.from(document.querySelectorAll('div[matlistitemtitle].text-menu-parent, div[matlistitemtitle].text-menu, div[matlistitemtitle].mat-mdc-list-item-title, .mat-mdc-list-item-title'));

for (const node of nodes) {
  const text = normalize(node.textContent || node.innerText || '');
  const style = window.getComputedStyle(node);
  const visible = style && style.visibility !== 'hidden' && style.display !== 'none' && node.offsetParent !== null;
  if (!visible || text !== wanted) continue;

    const candidates = [
        node,
        node.closest('a, button, [role="menuitem"], [role="listitem"], mat-list-item, .mat-mdc-list-item, .mdc-list-item, li, div')
    ].filter(Boolean);

    for (const target of candidates) {
        try { target.scrollIntoView({ block: 'center' }); } catch (e) {}
        try { target.focus(); } catch (e) {}

        ['pointerdown','mousedown','pointerup','mouseup','click'].forEach((type) => {
            target.dispatchEvent(new MouseEvent(type, { bubbles: true, cancelable: true, view: window }));
        });

        try { target.click(); } catch (e) {}
        return true;
    }
}

return false;
"""
        )
        return bool(clicked)
    except Exception:
        return False


def _perform_sne_click_attempt(driver, target, description: str, mode: str) -> bool:
    if target is None:
        return False

    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", target)
    except Exception:
        pass

    try:
        if mode == "selenium":
            if _click_ingresar_button(driver, target):
                logging.info("Clic real ejecutado sobre %s.", description)
                return True
        elif mode == "js-click":
            clicked = driver.execute_script(
                """
const el = arguments[0];
if (!el) return false;
el.click();
return true;
""",
                target,
            )
            if clicked:
                logging.info("Clic JS ejecutado sobre %s.", description)
                return True
        elif mode == "js-events":
            clicked = driver.execute_script(
                """
const el = arguments[0];
if (!el) return false;
['pointerdown','mousedown','pointerup','mouseup','click'].forEach((type) => {
  el.dispatchEvent(new MouseEvent(type, { bubbles: true, cancelable: true, view: window }));
});
return true;
""",
                target,
            )
            if clicked:
                logging.info("Eventos de click disparados sobre %s.", description)
                return True
        elif mode == "keyboard" and Keys is not None:
            target.send_keys(Keys.ENTER)
            logging.info("Enter enviado sobre %s.", description)
            return True
    except Exception:
        return False

    return False


def _switch_to_sne_window_if_any(driver, cfg: AuthConfig) -> bool:
    """Busca entre todas las ventanas la que pertenece al SNE y cambia el foco."""
    target_host = "notificaciones.osinergmin.gob.pe"
    expected_url = _normalize_text(cfg.sne_target_url)

    for handle in driver.window_handles:
        try:
            driver.switch_to.window(handle)
            current_url = driver.current_url or ""
        except Exception:
            continue

        current_url_norm = _normalize_text(current_url)
        if target_host in current_url.lower() or (expected_url and expected_url in current_url_norm):
            return True

    return False


def _wait_for_sne_navigation_after_click(driver, cfg: AuthConfig, windows_before: list[str], url_before: str) -> bool:
    """Confirma que el clic produjo navegacion real al SNE."""
    end_time = time.time() + cfg.timeout
    while time.time() < end_time:
        try:
            windows_now = driver.window_handles
        except Exception:
            return False

        if len(windows_now) > len(windows_before) and _switch_to_sne_window_if_any(driver, cfg):
            return True

        try:
            current_url = driver.current_url or ""
        except Exception:
            return False
        if _normalize_text(cfg.sne_target_url) in _normalize_text(current_url):
            return True

        if current_url and current_url != url_before and "notificaciones.osinergmin.gob.pe" in current_url.lower():
            return True

        time.sleep(0.25)

    return False


def _attempt_sne_click_navigation(driver, menu, cfg: AuthConfig, windows_before: list[str], url_before: str) -> bool:
    container, title = _resolve_sne_menu_targets(driver, menu)
    attempts = [
        (title, "el titulo de Casilla Electrónica del SNE", "selenium"),
        (container, "el contenedor del menu SNE", "selenium"),
        (title, "el titulo de Casilla Electrónica del SNE", "js-events"),
        (title, "el titulo de Casilla Electrónica del SNE", "js-click"),
        (container, "el contenedor del menu SNE", "js-click"),
        (container, "el contenedor del menu SNE", "js-events"),
        (title, "el titulo de Casilla Electrónica del SNE", "keyboard"),
        (container, "el contenedor del menu SNE", "keyboard"),
    ]

    for target, description, mode in attempts:
        if not _perform_sne_click_attempt(driver, target, description, mode):
            continue

        if _wait_for_sne_navigation_after_click(driver, cfg, windows_before, url_before):
            return True

    return False


def _fast_click_sne_menu_via_js(driver, cfg: AuthConfig) -> bool:
    """Clic rapido JS directo sobre mat-list-item padre del titulo Casilla SNE."""
    try:
        return bool(driver.execute_script(
            """
const normalize = (v) => (v || '').normalize('NFD').replace(/[\u0300-\u036f]/g, '').trim().toLowerCase();
const wanted = normalize('Casilla Electrónica del SNE');
const nodes = Array.from(document.querySelectorAll(
    'div[matlistitemtitle].text-menu-parent, div[matlistitemtitle].text-menu, div[matlistitemtitle]'
));
for (const node of nodes) {
    if (normalize(node.textContent || '') !== wanted) continue;
    const item = node.closest('mat-list-item, .mdc-list-item, [role="listitem"]') || node;
    item.scrollIntoView({ block: 'center' });
    ['pointerdown','mousedown','pointerup','mouseup','click'].forEach((t) => {
        item.dispatchEvent(new MouseEvent(t, { bubbles: true, cancelable: true, view: window }));
    });
    item.click();
    return true;
}
return false;
"""
        ))
    except Exception:
        return False


def _click_sne_menu_and_switch_window(driver, cfg: AuthConfig) -> bool:
    """Hace clic en Casilla Electrónica del SNE y cambia a la ventana nueva si se abre."""
    windows_before = driver.window_handles
    url_before = driver.current_url or ""

    # Esperar a que Angular renderice el elemento antes de intentar el clic.
    sne_selectors = [
        (By.CSS_SELECTOR, "div[matlistitemtitle].text-menu-parent"),
        (By.CSS_SELECTOR, "div[matlistitemtitle].text-menu"),
        (By.XPATH, "//div[@matlistitemtitle and contains(normalize-space(),'SNE')]"),
    ]
    for by, val in sne_selectors:
        try:
            WebDriverWait(driver, cfg.timeout).until(EC.presence_of_element_located((by, val)))
            logging.info("Elemento del menu SNE detectado en el DOM.")
            break
        except Exception:
            continue

    clicked = False
    deadline = time.time() + max(20, cfg.timeout * 2)
    while time.time() < deadline and not clicked:
        # Ruta rapida: JS directo sobre el elemento exacto del DOM.
        if _fast_click_sne_menu_via_js(driver, cfg):
            logging.info("Clic JS rapido ejecutado sobre Casilla Electronica del SNE.")
            clicked = _wait_for_sne_navigation_after_click(driver, cfg, windows_before, url_before)
            if clicked:
                break

        if not clicked:
            try:
                menu = _find_sne_menu(driver, cfg, per_selector_timeout=2)
            except Exception:
                menu = None

            if menu is not None:
                clicked = _attempt_sne_click_navigation(driver, menu, cfg, windows_before, url_before)

        if not clicked and _force_click_sne_menu_via_js(driver):
            clicked = _wait_for_sne_navigation_after_click(driver, cfg, windows_before, url_before)

        if not clicked:
            time.sleep(0.5)

    if not clicked:
        if cfg.require_sne_click_navigation:
            logging.warning("No se pudo hacer clic en el menu 'Casilla Electrónica del SNE'.")
            return False
        logging.warning("No se pudo hacer clic en el menu 'Casilla Electrónica del SNE'; se usara navegacion directa.")

    if not clicked and cfg.sne_target_url:
        logging.info("Abriendo el link objetivo del SNE automaticamente.")
        try:
            driver.get(cfg.sne_target_url)
        except Exception as exc:
            logging.warning("No se pudo abrir la URL objetivo del SNE: %s", exc)
            return False

    if _wait_for_sne_home(driver, cfg):
        logging.info("El clic abrio el SNE en la URL esperada y con el contenido de bandeja.")
    else:
        logging.warning("El clic abrio una vista relacionada, pero no se pudo validar la pantalla esperada del SNE.")
    return True


def _element_exists(driver, css_selector: str) -> bool:
    if not css_selector:
        return False
    try:
        return len(driver.find_elements(By.CSS_SELECTOR, css_selector)) > 0
    except Exception:
        return False


def _has_captcha(driver) -> bool:
    return (
        _element_exists(driver, "iframe[src*='recaptcha']")
        or _element_exists(driver, "iframe[title*='recaptcha' i]")
        or "captcha" in (driver.page_source or "").lower()
    )


def _wait_until_submit_enabled(driver, cfg: AuthConfig) -> bool:
    btn = _find_ingresar_button(driver, cfg, 2)
    if btn is None:
        return False

    disabled = (btn.get_attribute("disabled") or "").strip().lower()
    aria_disabled = (btn.get_attribute("aria-disabled") or "").strip().lower()
    return disabled in {"", "false"} and aria_disabled != "true"


def _wait_for_captcha_resolution(driver, cfg: AuthConfig) -> bool:
    """Espera a que el captcha se resuelva en modo manual (navegador visible)."""

    end_time = time.time() + cfg.captcha_timeout
    while time.time() < end_time:
        token_nodes = driver.find_elements(By.CSS_SELECTOR, "textarea[name='g-recaptcha-response']")
        if token_nodes:
            token = (token_nodes[0].get_attribute("value") or "").strip()
            if token:
                return True

        if _wait_until_submit_enabled(driver, cfg):
            return True

        time.sleep(0.5)

    return False


def _wait_for_login_result(driver, cfg: AuthConfig) -> bool:
    """Espera señales de éxito/fallo sin depender solo del cambio de URL."""

    def _condition(drv) -> bool:
        current_url = drv.current_url or ""

        # Exito 1: URL cambia a una ruta diferente de login.
        if _is_authenticated_by_url(current_url, cfg.success_url_keyword):
            return True

        # Exito 2: aparece un selector conocido de pagina autenticada.
        if cfg.post_login_selector and _element_exists(drv, cfg.post_login_selector):
            return True

        # Exito 3: desaparece el campo de password del login.
        if not drv.find_elements(By.ID, cfg.password_id):
            return True

        # Fallo temprano: aparece un mensaje/selector de error de autenticacion.
        if cfg.login_error_selector and _element_exists(drv, cfg.login_error_selector):
            return False

        return False

    try:
        WebDriverWait(driver, cfg.timeout).until(_condition)
    except TimeoutException:
        logging.warning("Timeout esperando resultado de login con Selenium.")

    current_url = driver.current_url or ""
    if _is_authenticated_by_url(current_url, cfg.success_url_keyword):
        return True

    if cfg.post_login_selector and _element_exists(driver, cfg.post_login_selector):
        return True

    # Si seguimos viendo el input de password, normalmente seguimos en login.
    return len(driver.find_elements(By.ID, cfg.password_id)) == 0


def _login_with_selenium(session: requests.Session, cfg: AuthConfig, username: str, password: str) -> bool:
    driver = _new_driver(cfg)
    direct_sne_url = "https://notificaciones.osinergmin.gob.pe/sne-web/pages/notificacion/inicio"
    try:
        driver.get(cfg.login_url)

        user_input = _find_first(
            driver,
            [
                (By.ID, cfg.username_id),
                (By.CSS_SELECTOR, "input[type='text']"),
                (By.CSS_SELECTOR, "input[autocomplete='username']"),
                (By.CSS_SELECTOR, "input[id*='user'], input[name*='user']"),
            ],
            cfg.timeout,
        )
        pass_input = _find_first(
            driver,
            [
                (By.ID, cfg.password_id),
                (By.CSS_SELECTOR, "input[type='password']"),
                (By.CSS_SELECTOR, "input[autocomplete='current-password']"),
            ],
            cfg.timeout,
        )

        if user_input is None or pass_input is None:
            raise RuntimeError("No se encontraron campos de usuario/contrasena en la pagina.")

        user_input.clear()
        user_input.send_keys(username)
        pass_input.clear()
        pass_input.send_keys(password)

        submit = _find_ingresar_button(driver, cfg, cfg.timeout)
        if submit is None:
            logging.warning("No se encontro boton inicialmente por selectores.")

        # Este portal usa captcha; forzamos flujo manual si se detecta.
        disabled = (submit.get_attribute("disabled") or "").strip().lower() if submit is not None else "<sin-boton>"
        has_captcha = _has_captcha(driver)
        logging.info("Selenium pre-submit: has_captcha=%s disabled=%s", has_captcha, disabled or "<vacio>")

        if has_captcha:
            if cfg.selenium_headless:
                raise RuntimeError(
                    "Se detecto captcha. Ejecuta con OSI_HEADLESS=false para resolverlo manualmente en el navegador."
                )

            print("Captcha detectado: resuelvelo en la ventana del navegador.")
            print("Esperando resolucion del captcha...")
            if not _wait_for_captcha_resolution(driver, cfg):
                raise RuntimeError("No se detecto captcha resuelto dentro del tiempo permitido.")

            logging.info("Captcha resuelto detectado. Ejecutando clic automatico en Ingresar...")

        if not _click_ingresar_with_retries(driver, cfg):
            raise RuntimeError("No se pudo hacer clic automático en el boton Ingresar.")

        if not _wait_for_login_result(driver, cfg):
            return False

        if cfg.open_sne_after_login:
            # El SNE es un dominio distinto; primero establecemos la sesion SSO
            # via el portal (necesario para que notificaciones.osinergmin.gob.pe
            # acepte la sesion), y luego navegamos a la URL exacta solicitada.
            _click_sne_menu_and_switch_window(driver, cfg)
            logging.info("Redirigiendo a la URL directa del SNE: %s", direct_sne_url)
            try:
                driver.get(direct_sne_url)
            except Exception as exc:
                logging.warning("No se pudo navegar a la URL directa del SNE: %s", exc)
                return False

            if not _wait_for_sne_home(driver, cfg):
                logging.warning("No se pudo validar la pantalla esperada del SNE tras navegar a la URL objetivo.")

            _apply_sne_filters(driver, cfg)

        # Limpia cualquier alerta remanente y copia cookies ANTES del sleep,
        # mientras la sesion Selenium sigue activa.
        _accept_browser_alert_if_present(driver, expected_text="No hay datos para exportar")
        try:
            for cookie in driver.get_cookies():
                session.cookies.set(cookie["name"], cookie["value"], domain=cookie.get("domain"), path=cookie.get("path", "/"))
        except Exception:
            pass

        if cfg.open_sne_after_login:
            # Deja el navegador abierto 1 minuto para que puedas ver el resultado.
            logging.info("Esperando 1 minuto para que puedas revisar los resultados en el navegador...")
            time.sleep(60)
            logging.info("Cerrando navegador.")

        return True
    finally:
        try:
            driver.quit()
        except Exception:
            pass


def login(session: requests.Session, cfg: AuthConfig, username: str, password: str) -> None:
    """Autentica contra OsiVirtual usando requests y, si es necesario, Selenium."""
    ok = False
    try:
        ok = _login_with_requests(session, cfg, username, password)
        if ok:
            logging.info("Login exitoso con requests.")
            return
        logging.info("Login por requests no concluyente; se intentara con Selenium.")
    except Exception as exc:
        logging.warning("Fallo login con requests: %s", exc)

    ok = _login_with_selenium(session, cfg, username, password)
    if not ok:
        raise RuntimeError("No se pudo autenticar con las credenciales proporcionadas.")

    logging.info("Login exitoso con Selenium.")


def main() -> None:
    _load_dotenv()

    parser = argparse.ArgumentParser(description="Prueba de autenticacion a OsiVirtual.")
    parser.add_argument("--username", default=os.getenv("OSI_USERNAME"), help="Usuario de OsiVirtual (env: OSI_USERNAME)")
    parser.add_argument("--password", default=os.getenv("OSI_PASSWORD"), help="Contrasena de OsiVirtual (env: OSI_PASSWORD)")
    parser.add_argument(
        "--login-url",
        default=os.getenv("OSI_LOGIN_URL", "https://osivirtual.osinergmin.gob.pe/autenticacion/acceso-sistema"),
        help="URL de login (env: OSI_LOGIN_URL)",
    )
    parser.add_argument(
        "--base-url",
        default=os.getenv("OSI_BASE_URL", "https://osivirtual.osinergmin.gob.pe"),
        help="URL base (env: OSI_BASE_URL)",
    )
    parser.add_argument(
        "--username-id",
        default=os.getenv("OSI_USERNAME_ID", "documentoIdentidad"),
        help="ID del campo usuario (env: OSI_USERNAME_ID)",
    )
    parser.add_argument(
        "--password-id",
        default=os.getenv("OSI_PASSWORD_ID", "contrasena"),
        help="ID del campo contrasena (env: OSI_PASSWORD_ID)",
    )
    parser.add_argument(
        "--login-form-selector",
        default=os.getenv("OSI_LOGIN_FORM_SELECTOR", "form"),
        help="Selector CSS del formulario (env: OSI_LOGIN_FORM_SELECTOR)",
    )
    parser.add_argument(
        "--login-submit-selector",
        default=os.getenv("OSI_LOGIN_SUBMIT_SELECTOR", "button[type='submit']"),
        help="Selector CSS del boton enviar (env: OSI_LOGIN_SUBMIT_SELECTOR)",
    )
    parser.add_argument(
        "--post-login-selector",
        default=os.getenv("OSI_POST_LOGIN_SELECTOR", ""),
        help="Selector CSS visible tras login exitoso (env: OSI_POST_LOGIN_SELECTOR)",
    )
    parser.add_argument(
        "--login-error-selector",
        default=os.getenv("OSI_LOGIN_ERROR_SELECTOR", ""),
        help="Selector CSS de error de login (env: OSI_LOGIN_ERROR_SELECTOR)",
    )
    parser.add_argument(
        "--success-url-keyword",
        default=os.getenv("OSI_SUCCESS_URL_KEYWORD", "autenticacion/acceso-sistema"),
        help="Texto que indica que sigue en login (env: OSI_SUCCESS_URL_KEYWORD)",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=int(os.getenv("OSI_TIMEOUT", "30")),
        help="Timeout en segundos (env: OSI_TIMEOUT)",
    )
    parser.add_argument(
        "--captcha-timeout",
        type=int,
        default=int(os.getenv("OSI_CAPTCHA_TIMEOUT", "180")),
        help="Tiempo maximo para resolver captcha (env: OSI_CAPTCHA_TIMEOUT)",
    )
    parser.add_argument(
        "--open-sne-after-login",
        action="store_true",
        default=_env_bool("OSI_OPEN_SNE_AFTER_LOGIN", True),
        help="Abre Casilla Electrónica del SNE tras login (env: OSI_OPEN_SNE_AFTER_LOGIN)",
    )
    parser.add_argument(
        "--require-sne-click-navigation",
        action="store_true",
        default=_env_bool("OSI_REQUIRE_SNE_CLICK_NAVIGATION", True),
        help="Exige que el SNE se abra por clic real y no por navegacion forzada (env: OSI_REQUIRE_SNE_CLICK_NAVIGATION)",
    )
    parser.add_argument(
        "--sne-menu-selector",
        default=os.getenv(
            "OSI_SNE_MENU_SELECTOR",
            "//div[@matlistitemtitle and contains(@class,'text-menu-parent') "
            "and contains(normalize-space(),'Casilla') and contains(normalize-space(),'SNE')]",
        ),
        help="XPath/CSS (XPath recomendado) del menu SNE (env: OSI_SNE_MENU_SELECTOR)",
    )
    parser.add_argument(
        "--sne-target-url",
        default=os.getenv(
            "OSI_SNE_TARGET_URL",
            "https://notificaciones.osinergmin.gob.pe/sne-web/pages/notificacion/inicio",
        ),
        help="URL objetivo de Casilla Electronica del SNE (env: OSI_SNE_TARGET_URL)",
    )
    parser.add_argument(
        "--sne-expected-text",
        default=os.getenv(
            "OSI_SNE_EXPECTED_TEXT",
            "Sistema de Notificaciones Electrónicas|Bandeja de Entrada",
        ),
        help="Textos esperados en pantalla SNE separados por | (env: OSI_SNE_EXPECTED_TEXT)",
    )
    parser.add_argument(
        "--fecha-notificacion-inicio",
        default=os.getenv("OSI_FECHA_NOTIFICACION_INICIO", ""),
        help="Fecha inicial para buscar en SNE con formato dd/mm/yyyy (env: OSI_FECHA_NOTIFICACION_INICIO)",
    )
    parser.add_argument(
        "--fecha-notificacion-fin",
        default=os.getenv("OSI_FECHA_NOTIFICACION_FIN", ""),
        help="Fecha final para buscar en SNE con formato dd/mm/yyyy (env: OSI_FECHA_NOTIFICACION_FIN)",
    )
    parser.add_argument(
        "--sne-fecha-inicio-id",
        default=os.getenv("OSI_SNE_FECHA_INICIO_ID", "fechaNotificacionInicio"),
        help="ID del campo fecha inicial del SNE (env: OSI_SNE_FECHA_INICIO_ID)",
    )
    parser.add_argument(
        "--sne-fecha-fin-id",
        default=os.getenv("OSI_SNE_FECHA_FIN_ID", "fechaNotificacionFin"),
        help="ID del campo fecha final del SNE (env: OSI_SNE_FECHA_FIN_ID)",
    )
    parser.add_argument(
        "--sne-leido-value",
        default=os.getenv("OSI_SNE_LEIDO_VALUE", ""),
        help="Valor del filtro Estado de lectura en SNE (ej.: ''=Todos, 'N'=No leida, 'L'=Leida) (env: OSI_SNE_LEIDO_VALUE)",
    )
    parser.add_argument(
        "--sne-buscar-button-id",
        default=os.getenv("OSI_SNE_BUSCAR_BUTTON_ID", "buscar-boton"),
        help="ID del boton Buscar en el SNE (env: OSI_SNE_BUSCAR_BUTTON_ID)",
    )
    parser.add_argument(
        "--sne-export-excel-selector",
        default=os.getenv(
            "OSI_SNE_EXPORT_EXCEL_SELECTOR",
            "//div[contains(@class,'ui-pg-div') and .//span[contains(@class,'ui-icon-arrowthickstop-1-s')] and contains(normalize-space(.), 'Exportar a Excel')]",
        ),
        help="XPath/CSS del boton Exportar a Excel (env: OSI_SNE_EXPORT_EXCEL_SELECTOR)",
    )
    parser.add_argument(
        "--download-dir",
        default=_resolve_default_download_dir(),
        help="Carpeta de descargas para exportaciones (env: OSI_DOWNLOAD_DIR)",
    )
    parser.add_argument(
        "--export-wait-seconds",
        type=int,
        default=int(os.getenv("OSI_EXPORT_WAIT_SECONDS", "40")),
        help="Segundos maximos para esperar el archivo Excel descargado (env: OSI_EXPORT_WAIT_SECONDS)",
    )
    parser.add_argument(
        "--target-notifications",
        default=os.getenv("OSI_TARGET_NOTIFICATIONS", ""),
        help="Lista de Nro. Notificacion separados por coma/espacio para procesar solo esos (env: OSI_TARGET_NOTIFICATIONS)",
    )
    parser.add_argument(
        "--incremental-only",
        action="store_true",
        default=_env_bool("OSI_INCREMENTAL_ONLY", True),
        help="Solo procesa notificaciones nuevas o pendientes (env: OSI_INCREMENTAL_ONLY)",
    )
    parser.add_argument(
        "--skip-existing-notifications",
        action="store_true",
        default=_env_bool("OSI_SKIP_EXISTING_NOTIFICATIONS", True),
        help="No vuelve a descargar notificaciones que ya tienen archivos (env: OSI_SKIP_EXISTING_NOTIFICATIONS)",
    )
    parser.add_argument(
        "--user-agent",
        default=os.getenv(
            "OSI_USER_AGENT",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        ),
        help="User-Agent HTTP (env: OSI_USER_AGENT)",
    )
    parser.add_argument(
        "--headless",
        action="store_true",
        default=_env_bool("OSI_HEADLESS", True),
        help="Ejecuta Selenium en modo headless (env: OSI_HEADLESS)",
    )
    parser.add_argument(
        "--auto-backfill-start",
        default=os.getenv("OSI_AUTO_BACKFILL_START", ""),
        help="Fecha inicio YYYY-MM-DD para procesar por dias hasta --auto-backfill-end o hoy (env: OSI_AUTO_BACKFILL_START)",
    )
    parser.add_argument(
        "--auto-backfill-end",
        default=os.getenv("OSI_AUTO_BACKFILL_END", ""),
        help="Fecha fin YYYY-MM-DD para backfill por dias (env: OSI_AUTO_BACKFILL_END)",
    )
    parser.add_argument(
        "--daemon",
        action="store_true",
        default=_env_bool("OSI_DAEMON", False),
        help="Ejecuta en bucle automatico mientras la maquina este encendida (env: OSI_DAEMON)",
    )
    parser.add_argument(
        "--daemon-interval-minutes",
        type=int,
        default=int(os.getenv("OSI_DAEMON_INTERVAL_MINUTES", "60")),
        help="Minutos entre ciclos en modo daemon (env: OSI_DAEMON_INTERVAL_MINUTES)",
    )
    args = parser.parse_args()

    if not args.username or not args.password:
        raise RuntimeError("Faltan credenciales. Define OSI_USERNAME y OSI_PASSWORD o pasalas por CLI.")

    logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")

    cfg = AuthConfig(
        login_url=args.login_url,
        base_url=args.base_url,
        username_id=args.username_id,
        password_id=args.password_id,
        login_form_selector=args.login_form_selector,
        login_submit_selector=args.login_submit_selector,
        post_login_selector=args.post_login_selector,
        login_error_selector=args.login_error_selector,
        success_url_keyword=args.success_url_keyword,
        selenium_headless=args.headless,
        timeout=args.timeout,
        captcha_timeout=args.captcha_timeout,
        open_sne_after_login=args.open_sne_after_login,
        require_sne_click_navigation=args.require_sne_click_navigation,
        sne_menu_selector=args.sne_menu_selector,
        sne_target_url=args.sne_target_url,
        sne_expected_text=args.sne_expected_text,
        fecha_notificacion_inicio=args.fecha_notificacion_inicio,
        fecha_notificacion_fin=args.fecha_notificacion_fin,
        sne_fecha_inicio_id=args.sne_fecha_inicio_id,
        sne_fecha_fin_id=args.sne_fecha_fin_id,
        sne_leido_value=args.sne_leido_value,
        sne_buscar_button_id=args.sne_buscar_button_id,
        sne_export_excel_selector=args.sne_export_excel_selector,
        download_dir=args.download_dir,
        export_wait_seconds=args.export_wait_seconds,
        target_notifications=_parse_target_notifications(args.target_notifications),
        incremental_only=args.incremental_only,
        skip_existing_notifications=args.skip_existing_notifications,
        user_agent=args.user_agent,
    )

    session = requests.Session()
    session.headers.update({"User-Agent": cfg.user_agent})

    def _run_for_single_day(day: datetime) -> None:
        day_ddmmyyyy = day.strftime("%d/%m/%Y")
        day_cfg = replace(
            cfg,
            fecha_notificacion_inicio=day_ddmmyyyy,
            fecha_notificacion_fin=day_ddmmyyyy,
            incremental_only=True,
            skip_existing_notifications=True,
        )
        logging.info("Procesando dia %s en modo automatico.", day.strftime("%Y-%m-%d"))
        login(session, day_cfg, args.username, args.password)

    def _run_backfill_range(start_day: datetime, end_day: datetime) -> None:
        current = start_day
        while current <= end_day:
            _run_for_single_day(current)
            current = current + timedelta(days=1)

    try:
        backfill_start = _parse_iso_date(args.auto_backfill_start)
        backfill_end = _parse_iso_date(args.auto_backfill_end) or datetime.now()

        if backfill_start is not None:
            if backfill_start > backfill_end:
                raise RuntimeError("--auto-backfill-start no puede ser mayor que --auto-backfill-end.")
            _run_backfill_range(backfill_start, backfill_end)
            print("Backfill completado correctamente.")

        if args.daemon:
            interval_seconds = max(60, args.daemon_interval_minutes * 60)
            logging.info("Modo daemon activo. Intervalo: %s minutos.", interval_seconds // 60)
            while True:
                today = datetime.now()
                try:
                    _run_for_single_day(today)
                    logging.info("Ciclo automatico completado para %s.", today.strftime("%Y-%m-%d"))
                except Exception as exc:
                    logging.warning("Fallo en ciclo automatico: %s", exc)
                time.sleep(interval_seconds)

        if backfill_start is None and not args.daemon:
            login(session, cfg, args.username, args.password)
            print("Autenticacion completada correctamente.")
    except KeyboardInterrupt:
        print("Ejecucion interrumpida por el usuario.")


if __name__ == "__main__":
    main()
